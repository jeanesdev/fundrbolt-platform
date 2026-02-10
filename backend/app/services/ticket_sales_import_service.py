"""Service for ticket sales bulk import."""

from __future__ import annotations

import csv
import hashlib
import io
import json
from dataclasses import dataclass
from datetime import datetime
from decimal import Decimal, InvalidOperation
from typing import Any
from uuid import UUID

import magic
from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.ticket_management import (
    DiscountType,
    PromoCode,
    PromoCodeApplication,
    TicketPackage,
    TicketPurchase,
)
from app.models.ticket_sales_import import (
    ImportFormat,
    ImportStatus,
    IssueSeverity,
    TicketSalesImportBatch,
    TicketSalesImportIssue,
)
from app.schemas.ticket_sales_import import (
    ImportResult,
    ImportRowStatus,
    PreflightIssue,
    PreflightResult,
)

# Constants
MAX_IMPORT_ROWS = 5000
REQUIRED_CSV_HEADERS = [
    "ticket_type",
    "purchaser_name",
    "purchaser_email",
    "quantity",
    "total_amount",
    "purchase_date",
    "external_sale_id",
]
OPTIONAL_CSV_HEADERS = [
    "purchaser_phone",
    "promo_code",
    "discount_amount",
    "fee_amount",
    "payment_status",
    "notes",
]


@dataclass
class ParsedRow:
    """Parsed row from import file."""

    row_number: int
    data: dict[str, Any]


class TicketSalesImportError(Exception):
    """Base exception for ticket sales import errors."""

    pass


class TicketSalesImportService:
    """Bulk import service for ticket sales."""

    def __init__(self, db: AsyncSession) -> None:
        self.db = db

    async def preflight(
        self,
        event_id: UUID,
        file_bytes: bytes,
        filename: str,
        created_by: UUID,
    ) -> PreflightResult:
        """Validate uploaded file and return preflight results.

        Args:
            event_id: UUID of the event
            file_bytes: Raw bytes of the uploaded file
            filename: Original filename
            created_by: UUID of the user initiating the preflight

        Returns:
            PreflightResult with validation summary and issues

        Raises:
            TicketSalesImportError: If file format is unsupported or invalid
        """
        # Detect file format
        detected_format = self._detect_format(file_bytes, filename)

        # Parse file
        parsed_rows = self._parse_file(file_bytes, detected_format)

        # Check row count limit
        if len(parsed_rows) > MAX_IMPORT_ROWS:
            raise TicketSalesImportError(
                f"File contains {len(parsed_rows)} rows. Maximum allowed is {MAX_IMPORT_ROWS}."
            )

        # Fetch existing external_sale_ids for this event
        external_ids = [
            str(row.data.get("external_sale_id", "")).strip()
            for row in parsed_rows
            if row.data.get("external_sale_id")
        ]
        existing_ids = await self._fetch_existing_external_ids(event_id, external_ids)

        # Fetch ticket packages for validation
        ticket_packages = await self._fetch_ticket_packages(event_id)
        package_names = {pkg.name.lower() for pkg in ticket_packages}

        promo_codes = await self._fetch_promo_codes(event_id)
        promo_code_map = {promo.code.lower(): promo for promo in promo_codes}

        # Validate rows
        issues: list[PreflightIssue] = []
        warnings: list[PreflightIssue] = []
        valid_count = 0
        error_count = 0
        warning_count = 0

        for parsed_row in parsed_rows:
            row_issues = self._validate_row(
                parsed_row, package_names, existing_ids, external_ids, promo_code_map
            )
            if row_issues:
                for issue in row_issues:
                    if issue.severity == IssueSeverity.ERROR:
                        error_count += 1
                        issues.append(issue)
                    else:
                        warning_count += 1
                        warnings.append(issue)
            else:
                valid_count += 1

        # Calculate checksum
        checksum = hashlib.sha256(file_bytes).hexdigest()

        # Create import batch record
        batch = TicketSalesImportBatch(
            event_id=event_id,
            created_by=created_by,
            status=ImportStatus.PREFLIGHTED,
            source_filename=filename,
            source_format=detected_format,
            row_count=len(parsed_rows),
            valid_count=valid_count,
            error_count=error_count,
            warning_count=warning_count,
            preflight_checksum=checksum,
        )
        self.db.add(batch)
        await self.db.flush()

        # Create issue records
        for issue in issues + warnings:
            issue_record = TicketSalesImportIssue(
                batch_id=batch.id,
                row_number=issue.row_number,
                field_name=issue.field_name,
                severity=IssueSeverity(issue.severity.value),
                message=issue.message,
                raw_value=issue.raw_value,
            )
            self.db.add(issue_record)

        return PreflightResult(
            preflight_id=str(batch.id),
            detected_format=detected_format.value,
            total_rows=len(parsed_rows),
            valid_rows=valid_count,
            error_rows=error_count,
            warning_rows=warning_count,
            issues=issues,
            warnings=warnings,
        )

    async def commit_import(
        self, event_id: UUID, preflight_id: UUID, file_bytes: bytes, user_id: UUID
    ) -> ImportResult:
        """Execute import using preflight results.

        Args:
            event_id: UUID of the event
            preflight_id: UUID of the preflight batch
            file_bytes: Raw bytes of the uploaded file (must match preflight checksum)
            user_id: UUID of the user initiating the import

        Returns:
            ImportResult with created, skipped, and failed counts

        Raises:
            TicketSalesImportError: If preflight not found or checksum mismatch
        """
        # Fetch preflight batch
        result = await self.db.execute(
            select(TicketSalesImportBatch).where(
                TicketSalesImportBatch.id == preflight_id,
                TicketSalesImportBatch.event_id == event_id,
            )
        )
        batch = result.scalar_one_or_none()
        if not batch:
            raise TicketSalesImportError("Preflight batch not found")

        # Verify checksum
        checksum = hashlib.sha256(file_bytes).hexdigest()
        if batch.preflight_checksum != checksum:
            raise TicketSalesImportError(
                "File has changed since preflight. Please re-run preflight."
            )

        # Check for errors
        if batch.error_count > 0:
            raise TicketSalesImportError(
                "Cannot import file with errors. Please fix errors and re-run preflight."
            )

        # Parse file
        parsed_rows = self._parse_file(file_bytes, batch.source_format)

        # Fetch existing external_sale_ids
        external_ids = [
            str(row.data.get("external_sale_id", "")).strip()
            for row in parsed_rows
            if row.data.get("external_sale_id")
        ]
        existing_ids = await self._fetch_existing_external_ids(event_id, external_ids)

        # Fetch ticket packages
        ticket_packages = await self._fetch_ticket_packages(event_id)
        package_map = {pkg.name.lower(): pkg for pkg in ticket_packages}

        promo_codes = await self._fetch_promo_codes(event_id)
        promo_code_map = {promo.code.lower(): promo for promo in promo_codes}

        # Create ticket purchases
        created_count = 0
        skipped_count = 0
        failed_count = 0
        warnings: list[PreflightIssue] = []

        for parsed_row in parsed_rows:
            try:
                external_id = str(parsed_row.data.get("external_sale_id", "")).strip()

                # Skip if external_sale_id already exists
                if external_id and external_id in existing_ids:
                    skipped_count += 1
                    warnings.append(
                        PreflightIssue(
                            row_number=parsed_row.row_number,
                            field_name="external_sale_id",
                            severity=IssueSeverity.WARNING,
                            message=f"Skipped: External sale ID '{external_id}' already exists",
                            raw_value=external_id,
                        )
                    )
                    continue

                # Get ticket package
                ticket_type = str(parsed_row.data.get("ticket_type", "")).strip()
                package = package_map.get(ticket_type.lower())
                if not package:
                    failed_count += 1
                    continue

                # Parse values
                quantity = int(parsed_row.data.get("quantity", 1))
                total_amount = Decimal(str(parsed_row.data.get("total_amount", 0)))
                purchase_date_str = str(parsed_row.data.get("purchase_date", ""))
                promo_code_raw = str(parsed_row.data.get("promo_code", "")).strip()
                discount_amount_raw = parsed_row.data.get("discount_amount")
                
                # Parse purchase date
                try:
                    purchase_date = datetime.fromisoformat(purchase_date_str)
                except ValueError:
                    # Try common date formats
                    from dateutil import parser as date_parser
                    purchase_date = date_parser.parse(purchase_date_str)

                # Create ticket purchase
                purchase = TicketPurchase(
                    event_id=event_id,
                    ticket_package_id=package.id,
                    user_id=user_id,
                    quantity=quantity,
                    total_price=total_amount,
                    payment_status="completed",  # Default to completed for imports
                    external_sale_id=external_id if external_id else None,
                    purchaser_name=str(parsed_row.data.get("purchaser_name", "")).strip()
                    or None,
                    purchaser_email=str(parsed_row.data.get("purchaser_email", "")).strip()
                    or None,
                    purchaser_phone=str(parsed_row.data.get("purchaser_phone", "")).strip()
                    or None,
                    notes=str(parsed_row.data.get("notes", "")).strip() or None,
                    purchased_at=purchase_date,
                )
                self.db.add(purchase)
                created_count += 1

                if promo_code_raw:
                    promo = promo_code_map.get(promo_code_raw.lower())
                    if not self._is_promo_code_valid(promo):
                        warnings.append(
                            PreflightIssue(
                                row_number=parsed_row.row_number,
                                field_name="promo_code",
                                severity=IssueSeverity.WARNING,
                                message=f"Promo code '{promo_code_raw}' is not valid",
                                raw_value=promo_code_raw,
                            )
                        )
                    else:
                        discount_amount = self._resolve_discount_amount(
                            total_amount, promo, discount_amount_raw
                        )
                        promo_application = PromoCodeApplication(
                            promo_code_id=promo.id,
                            ticket_purchase_id=purchase.id,
                            discount_amount=discount_amount,
                        )
                        self.db.add(promo_application)
                        promo.used_count += 1

                # Update sold_count on package
                package.sold_count += quantity

            except Exception as e:
                failed_count += 1
                warnings.append(
                    PreflightIssue(
                        row_number=parsed_row.row_number,
                        field_name=None,
                        severity=IssueSeverity.WARNING,
                        message=f"Failed to import: {str(e)}",
                    )
                )

        # Update batch status
        batch.status = ImportStatus.IMPORTED
        batch.created_by = user_id

        await self.db.flush()

        return ImportResult(
            batch_id=str(batch.id),
            created_rows=created_count,
            skipped_rows=skipped_count,
            failed_rows=failed_count,
            warnings=warnings,
        )

    def _detect_format(self, file_bytes: bytes, filename: str) -> ImportFormat:
        """Detect file format from bytes and filename."""
        mime_type = magic.from_buffer(file_bytes, mime=True)

        if mime_type == "text/csv" or filename.lower().endswith(".csv"):
            return ImportFormat.CSV
        elif mime_type == "application/json" or filename.lower().endswith(".json"):
            return ImportFormat.JSON
        elif mime_type in [
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "application/vnd.ms-excel",
        ] or filename.lower().endswith((".xlsx", ".xls")):
            return ImportFormat.XLSX
        else:
            raise TicketSalesImportError(
                f"Unsupported file format. Please upload a CSV, JSON, or Excel file. Detected: {mime_type}"
            )

    def _parse_file(self, file_bytes: bytes, file_format: ImportFormat) -> list[ParsedRow]:
        """Parse file based on format."""
        if file_format == ImportFormat.CSV:
            return self._parse_csv(file_bytes)
        elif file_format == ImportFormat.JSON:
            return self._parse_json(file_bytes)
        elif file_format == ImportFormat.XLSX:
            return self._parse_xlsx(file_bytes)
        else:
            raise TicketSalesImportError(f"Unsupported format: {file_format}")

    def _parse_csv(self, file_bytes: bytes) -> list[ParsedRow]:
        """Parse CSV file."""
        try:
            text = file_bytes.decode("utf-8")
            reader = csv.DictReader(io.StringIO(text))
            rows = []
            for idx, row in enumerate(reader, start=2):  # Start at 2 (header is row 1)
                rows.append(ParsedRow(row_number=idx, data=dict(row)))
            return rows
        except Exception as e:
            raise TicketSalesImportError(f"Failed to parse CSV: {str(e)}")

    def _parse_json(self, file_bytes: bytes) -> list[ParsedRow]:
        """Parse JSON file."""
        try:
            text = file_bytes.decode("utf-8")
            data = json.loads(text)
            if not isinstance(data, list):
                raise TicketSalesImportError("JSON must contain an array of ticket sales")
            rows = []
            for idx, item in enumerate(data, start=1):
                rows.append(ParsedRow(row_number=idx, data=item))
            return rows
        except json.JSONDecodeError as e:
            raise TicketSalesImportError(f"Invalid JSON: {str(e)}")
        except Exception as e:
            raise TicketSalesImportError(f"Failed to parse JSON: {str(e)}")

    def _parse_xlsx(self, file_bytes: bytes) -> list[ParsedRow]:
        """Parse Excel file."""
        try:
            workbook = load_workbook(io.BytesIO(file_bytes))
            sheet = workbook.active
            if sheet is None:
                raise TicketSalesImportError("Excel file has no active sheet")

            # Get headers from first row
            headers = [cell.value for cell in sheet[1]]

            rows = []
            for idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                data = {headers[i]: row[i] for i in range(len(headers)) if i < len(row)}
                rows.append(ParsedRow(row_number=idx, data=data))
            return rows
        except Exception as e:
            raise TicketSalesImportError(f"Failed to parse Excel: {str(e)}")

    def _validate_row(
        self,
        parsed_row: ParsedRow,
        package_names: set[str],
        existing_ids: set[str],
        all_ids_in_file: list[str],
        promo_code_map: dict[str, PromoCode],
    ) -> list[PreflightIssue]:
        """Validate a single row and return issues."""
        issues: list[PreflightIssue] = []
        data = parsed_row.data
        row_num = parsed_row.row_number

        # Check required fields
        for field in REQUIRED_CSV_HEADERS:
            value = data.get(field)
            if value is None or str(value).strip() == "":
                issues.append(
                    PreflightIssue(
                        row_number=row_num,
                        field_name=field,
                        severity=IssueSeverity.ERROR,
                        message=f"Missing required field: {field}",
                    )
                )

        # Validate ticket_type exists
        ticket_type = str(data.get("ticket_type", "")).strip()
        if ticket_type and ticket_type.lower() not in package_names:
            issues.append(
                PreflightIssue(
                    row_number=row_num,
                    field_name="ticket_type",
                    severity=IssueSeverity.ERROR,
                    message=f"Ticket type '{ticket_type}' not found for this event",
                    raw_value=ticket_type,
                )
            )

        # Validate quantity
        try:
            qty = int(data.get("quantity", 0))
            if qty <= 0:
                issues.append(
                    PreflightIssue(
                        row_number=row_num,
                        field_name="quantity",
                        severity=IssueSeverity.ERROR,
                        message="Quantity must be a positive integer",
                        raw_value=str(data.get("quantity")),
                    )
                )
        except (ValueError, TypeError):
            issues.append(
                PreflightIssue(
                    row_number=row_num,
                    field_name="quantity",
                    severity=IssueSeverity.ERROR,
                    message="Quantity must be a valid integer",
                    raw_value=str(data.get("quantity")),
                )
            )

        # Validate total_amount
        try:
            amount = Decimal(str(data.get("total_amount", 0)))
            if amount < 0:
                issues.append(
                    PreflightIssue(
                        row_number=row_num,
                        field_name="total_amount",
                        severity=IssueSeverity.ERROR,
                        message="Total amount must be non-negative",
                        raw_value=str(data.get("total_amount")),
                    )
                )
        except (InvalidOperation, ValueError, TypeError):
            issues.append(
                PreflightIssue(
                    row_number=row_num,
                    field_name="total_amount",
                    severity=IssueSeverity.ERROR,
                    message="Total amount must be a valid number",
                    raw_value=str(data.get("total_amount")),
                )
            )

        # Check for duplicate external_sale_id in file
        external_id = str(data.get("external_sale_id", "")).strip()
        if external_id:
            count_in_file = all_ids_in_file.count(external_id)
            if count_in_file > 1:
                issues.append(
                    PreflightIssue(
                        row_number=row_num,
                        field_name="external_sale_id",
                        severity=IssueSeverity.ERROR,
                        message=f"Duplicate external_sale_id '{external_id}' found in file",
                        raw_value=external_id,
                    )
                )

            # Check if external_sale_id already exists in database
            if external_id in existing_ids:
                issues.append(
                    PreflightIssue(
                        row_number=row_num,
                        field_name="external_sale_id",
                        severity=IssueSeverity.WARNING,
                        message=f"External sale ID '{external_id}' already exists and will be skipped",
                        raw_value=external_id,
                    )
                )

        promo_code_raw = str(data.get("promo_code", "")).strip()
        if promo_code_raw:
            promo = promo_code_map.get(promo_code_raw.lower())
            if not self._is_promo_code_valid(promo):
                issues.append(
                    PreflightIssue(
                        row_number=row_num,
                        field_name="promo_code",
                        severity=IssueSeverity.ERROR,
                        message=f"Promo code '{promo_code_raw}' is not valid for this event",
                        raw_value=promo_code_raw,
                    )
                )

        discount_raw = data.get("discount_amount")
        if discount_raw not in (None, ""):
            try:
                discount_value = Decimal(str(discount_raw))
                if discount_value < 0:
                    raise InvalidOperation
                try:
                    total_amount = Decimal(str(data.get("total_amount", 0)))
                    if discount_value > total_amount:
                        issues.append(
                            PreflightIssue(
                                row_number=row_num,
                                field_name="discount_amount",
                                severity=IssueSeverity.ERROR,
                                message="Discount amount cannot exceed total amount",
                                raw_value=str(discount_raw),
                            )
                        )
                except (InvalidOperation, ValueError, TypeError):
                    pass
            except (InvalidOperation, ValueError, TypeError):
                issues.append(
                    PreflightIssue(
                        row_number=row_num,
                        field_name="discount_amount",
                        severity=IssueSeverity.ERROR,
                        message="Discount amount must be a valid non-negative number",
                        raw_value=str(discount_raw),
                    )
                )

        return issues

    async def _fetch_existing_external_ids(
        self, event_id: UUID, external_ids: list[str]
    ) -> set[str]:
        """Fetch external sale IDs that already exist for this event."""
        if not external_ids:
            return set()

        result = await self.db.execute(
            select(TicketPurchase.external_sale_id).where(
                TicketPurchase.event_id == event_id,
                TicketPurchase.external_sale_id.in_(external_ids),
            )
        )
        existing = result.scalars().all()
        return {str(ext_id) for ext_id in existing if ext_id}

    async def _fetch_ticket_packages(self, event_id: UUID) -> list[TicketPackage]:
        """Fetch all ticket packages for the event."""
        result = await self.db.execute(
            select(TicketPackage).where(TicketPackage.event_id == event_id)
        )
        return list(result.scalars().all())

    async def _fetch_promo_codes(self, event_id: UUID) -> list[PromoCode]:
        """Fetch all promo codes for the event."""
        result = await self.db.execute(
            select(PromoCode).where(PromoCode.event_id == event_id)
        )
        return list(result.scalars().all())

    @staticmethod
    def _is_promo_code_valid(promo: PromoCode | None) -> bool:
        """Validate promo code status and date range."""
        if not promo:
            return False
        if not promo.is_active:
            return False
        now = datetime.utcnow()
        if promo.valid_from and now < promo.valid_from:
            return False
        if promo.valid_until and now > promo.valid_until:
            return False
        if promo.max_uses and promo.used_count >= promo.max_uses:
            return False
        return True

    @staticmethod
    def _resolve_discount_amount(
        total_amount: Decimal, promo: PromoCode, discount_raw: Any
    ) -> Decimal:
        """Resolve discount amount from explicit value or promo configuration."""
        if discount_raw not in (None, ""):
            try:
                return Decimal(str(discount_raw))
            except (InvalidOperation, ValueError, TypeError):
                return Decimal("0")
        if promo.discount_type == DiscountType.PERCENTAGE:
            return (total_amount * promo.discount_value) / Decimal("100")
        return min(total_amount, promo.discount_value)
