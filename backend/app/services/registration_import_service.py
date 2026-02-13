"""Service for registration bulk import."""

from __future__ import annotations

import base64
import csv
import io
import json
import uuid
from collections import Counter
from dataclasses import dataclass
from datetime import date, datetime
from typing import Any
from uuid import UUID

from openpyxl import load_workbook
from sqlalchemy import func, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.security import hash_password
from app.models.base import Base
from app.models.event import FoodOption
from app.models.event_registration import EventRegistration, RegistrationStatus
from app.models.meal_selection import MealSelection
from app.models.registration_guest import RegistrationGuest
from app.models.ticket_management import TicketPurchase
from app.models.user import User
from app.schemas.registration_import import (
    ImportErrorReportFormat,
    ImportErrorReportRequest,
    ImportErrorReportResponse,
    ImportReport,
    ImportRowResult,
    ImportRowStatus,
    ValidationIssue,
    ValidationIssueSeverity,
)

MAX_IMPORT_ROWS = 5000

REQUIRED_HEADERS = [
    "registrant_name",
    "registrant_email",
    "registration_date",
    "quantity",
    "external_registration_id",
]

OPTIONAL_HEADERS = [
    "event_id",
    "registrant_phone",
    "notes",
    "bidder_number",
    "table_number",
    "guest_count",
    "guest_of_email",
    "food_option",
    "ticket_purchase_id",
    "ticket_purchaser_email",
    "ticket_purchase_date",
]


@dataclass
class ParsedRow:
    row_number: int
    data: dict[str, Any]


class RegistrationImportError(Exception):
    """Base exception for registration import errors."""


class RegistrationImportService:
    """Bulk import service for event registrations."""

    def __init__(self, db: AsyncSession) -> None:
        self.db = db

    async def preflight(self, event_id: UUID, file_bytes: bytes, filename: str) -> ImportReport:
        """Run preflight validation without creating records."""
        file_type = self._detect_file_type(filename)
        parsed_rows = self._parse_file(file_bytes, file_type)

        # Fetch existing external IDs
        external_ids = [
            str(value).strip()
            for row in parsed_rows
            if (value := row.data.get("external_registration_id")) not in (None, "")
        ]
        existing_ids = await self._fetch_existing_external_ids(event_id, external_ids)

        # Validate all rows
        results = await self._validate_rows(event_id, parsed_rows, existing_ids)

        return self._build_report(results, file_type)

    async def commit(
        self, event_id: UUID, file_bytes: bytes, filename: str, user_id: UUID
    ) -> ImportReport:
        """Execute import and create registration records."""
        file_type = self._detect_file_type(filename)
        parsed_rows = self._parse_file(file_bytes, file_type)
        row_lookup = {row.row_number: row for row in parsed_rows}

        guest_rows_by_parent_email: dict[str, list[ParsedRow]] = {}
        guest_parent_emails: set[str] = set()
        for row in parsed_rows:
            guest_of_email = self._normalize_email(row.data.get("guest_of_email"))
            if guest_of_email:
                guest_rows_by_parent_email.setdefault(guest_of_email, []).append(row)
                guest_parent_emails.add(guest_of_email)

        # Fetch existing external IDs
        external_ids = [
            str(value).strip()
            for row in parsed_rows
            if (value := row.data.get("external_registration_id")) not in (None, "")
        ]
        existing_ids = await self._fetch_existing_external_ids(event_id, external_ids)

        existing_parent_by_email = await self._fetch_registrations_by_email(
            event_id, guest_parent_emails
        )
        food_option_name_index, food_option_id_index = await self._fetch_food_option_index(event_id)

        # Validate and then create records
        validation_results = await self._validate_rows(event_id, parsed_rows, existing_ids)

        results = []
        created_parent_by_email: dict[str, EventRegistration] = {}
        guest_rows_created: set[int] = set()

        for result in validation_results:
            row = row_lookup[result.row_number]
            is_guest_row = bool(self._normalize_email(row.data.get("guest_of_email")))
            if is_guest_row:
                continue

            if result.status == ImportRowStatus.ERROR:
                results.append(result)
                continue

            if result.status == ImportRowStatus.SKIPPED:
                results.append(result)
                continue

            try:
                parent_email = self._normalize_email(row.data.get("registrant_email"))
                guest_rows = [g.data for g in guest_rows_by_parent_email.get(parent_email, [])]
                created_registration = await self._create_registration(
                    event_id=event_id,
                    row=row.data,
                    user_id=user_id,
                    guest_rows=guest_rows,
                    food_option_name_index=food_option_name_index,
                    food_option_id_index=food_option_id_index,
                )
                if created_registration:
                    created_parent_by_email[parent_email] = created_registration
                    for guest_row in guest_rows_by_parent_email.get(parent_email, []):
                        guest_rows_created.add(guest_row.row_number)
                    results.append(result)
                else:
                    results.append(
                        ImportRowResult(
                            row_number=result.row_number,
                            external_id=result.external_id,
                            registrant_name=result.registrant_name,
                            registrant_email=result.registrant_email,
                            status=ImportRowStatus.SKIPPED,
                            message="Skipped: registration already exists",
                            issues=result.issues,
                        )
                    )
                await self.db.commit()
            except Exception as exc:
                await self.db.rollback()
                results.append(
                    ImportRowResult(
                        row_number=result.row_number,
                        external_id=result.external_id,
                        registrant_name=result.registrant_name,
                        registrant_email=result.registrant_email,
                        status=ImportRowStatus.ERROR,
                        message=f"Failed to create registration: {exc}",
                    )
                )

        for result in validation_results:
            row = row_lookup[result.row_number]
            guest_of_email = self._normalize_email(row.data.get("guest_of_email"))
            if not guest_of_email:
                continue

            if result.status == ImportRowStatus.ERROR:
                results.append(result)
                continue

            if result.row_number in guest_rows_created:
                results.append(result)
                continue

            parent_registration = created_parent_by_email.get(
                guest_of_email
            ) or existing_parent_by_email.get(guest_of_email)
            if not parent_registration:
                results.append(
                    ImportRowResult(
                        row_number=result.row_number,
                        external_id=result.external_id,
                        registrant_name=result.registrant_name,
                        registrant_email=result.registrant_email,
                        status=ImportRowStatus.ERROR,
                        message="Failed to create guest: parent registration not found",
                    )
                )
                continue

            try:
                await self._create_guest_registration(
                    parent_registration.id,
                    row.data,
                    food_option_name_index,
                    food_option_id_index,
                )
                await self.db.commit()
                results.append(result)
            except Exception as exc:
                await self.db.rollback()
                results.append(
                    ImportRowResult(
                        row_number=result.row_number,
                        external_id=result.external_id,
                        registrant_name=result.registrant_name,
                        registrant_email=result.registrant_email,
                        status=ImportRowStatus.ERROR,
                        message=f"Failed to create guest: {exc}",
                    )
                )

        return self._build_report(results, file_type)

    def _detect_file_type(self, filename: str) -> str:
        """Detect file type from filename."""
        filename_lower = filename.lower()
        if filename_lower.endswith(".json"):
            return "json"
        elif filename_lower.endswith(".csv"):
            return "csv"
        elif filename_lower.endswith((".xlsx", ".xls")):
            return "xlsx"
        else:
            raise RegistrationImportError(
                f"Unsupported file type. Must be JSON, CSV, or Excel (.xlsx): {filename}"
            )

    def _parse_file(self, file_bytes: bytes, file_type: str) -> list[ParsedRow]:
        """Parse file based on type."""
        if file_type == "json":
            return self._parse_json(file_bytes)
        elif file_type == "csv":
            return self._parse_csv(file_bytes)
        elif file_type == "xlsx":
            return self._parse_excel(file_bytes)
        else:
            raise RegistrationImportError(f"Unsupported file type: {file_type}")

    def _parse_json(self, file_bytes: bytes) -> list[ParsedRow]:
        """Parse JSON array of registration objects."""
        try:
            content = file_bytes.decode("utf-8")
            data = json.loads(content)
        except (UnicodeDecodeError, json.JSONDecodeError) as exc:
            raise RegistrationImportError(f"Invalid JSON file: {exc}") from exc

        if not isinstance(data, list):
            raise RegistrationImportError("JSON must contain an array of registration objects")

        if len(data) == 0:
            raise RegistrationImportError("JSON file contains no data")

        if len(data) > MAX_IMPORT_ROWS:
            raise RegistrationImportError(
                f"File exceeds maximum of {MAX_IMPORT_ROWS} rows (found {len(data)})"
            )

        parsed_rows: list[ParsedRow] = []
        for index, obj in enumerate(data, start=1):
            if not isinstance(obj, dict):
                continue
            row_data = {self._normalize_header(k): v for k, v in obj.items()}
            parsed_rows.append(ParsedRow(row_number=index, data=row_data))

        if not parsed_rows:
            raise RegistrationImportError("JSON file contains no valid data rows")

        return parsed_rows

    def _parse_csv(self, file_bytes: bytes) -> list[ParsedRow]:
        """Parse CSV file."""
        try:
            content = file_bytes.decode("utf-8-sig")
        except UnicodeDecodeError as exc:
            raise RegistrationImportError("CSV must be UTF-8 encoded") from exc

        reader = csv.reader(io.StringIO(content))
        header_row = next(reader, None)
        if not header_row:
            raise RegistrationImportError("CSV is empty")

        headers = [self._normalize_header(value) for value in header_row]
        self._validate_required_headers(headers)

        parsed_rows: list[ParsedRow] = []
        for index, values in enumerate(reader, start=2):
            if self._is_empty_row(values):
                continue
            if len(parsed_rows) >= MAX_IMPORT_ROWS:
                raise RegistrationImportError(f"CSV exceeds maximum of {MAX_IMPORT_ROWS} rows")

            row_data = {
                headers[i]: values[i] if i < len(values) else None for i in range(len(headers))
            }
            parsed_rows.append(ParsedRow(row_number=index, data=row_data))

        if not parsed_rows:
            raise RegistrationImportError("CSV contains no data rows")

        return parsed_rows

    def _parse_excel(self, file_bytes: bytes) -> list[ParsedRow]:
        """Parse Excel workbook."""
        try:
            workbook = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
        except Exception as exc:
            raise RegistrationImportError(f"Invalid Excel file: {exc}") from exc

        sheet = workbook.active
        row_iter = sheet.iter_rows(values_only=True)
        header_row = next(row_iter, None)
        if not header_row:
            raise RegistrationImportError("Excel workbook is empty")

        headers = [self._normalize_header(value) for value in header_row]
        self._validate_required_headers(headers)

        parsed_rows: list[ParsedRow] = []
        for index, values in enumerate(row_iter, start=2):
            if self._is_empty_row(values):
                continue
            if len(parsed_rows) >= MAX_IMPORT_ROWS:
                raise RegistrationImportError(f"Excel exceeds maximum of {MAX_IMPORT_ROWS} rows")

            row_data = {
                headers[i]: values[i] if i < len(values) else None for i in range(len(headers))
            }
            parsed_rows.append(ParsedRow(row_number=index, data=row_data))

        if not parsed_rows:
            raise RegistrationImportError("Excel workbook contains no data rows")

        return parsed_rows

    def _normalize_header(self, value: Any) -> str:
        """Normalize header name."""
        if value is None:
            return ""
        return str(value).strip().lower().replace(" ", "_").replace("-", "_")

    def _validate_required_headers(self, headers: list[str]) -> None:
        """Validate that all required headers are present."""
        missing = [h for h in REQUIRED_HEADERS if h not in headers]
        if missing:
            raise RegistrationImportError(f"Missing required columns: {', '.join(missing)}")

    def _is_empty_row(self, values: tuple[Any, ...] | list[Any]) -> bool:
        """Check if a row is empty."""
        return all(v is None or str(v).strip() == "" for v in values)

    async def _fetch_existing_external_ids(
        self, event_id: UUID, external_ids: list[str]
    ) -> set[str]:
        """Fetch existing external registration IDs for the event."""
        if not external_ids:
            return set()

        # TODO: Add external_registration_id field to EventRegistration model
        # For now, return empty set
        return set()

    async def _validate_rows(
        self, event_id: UUID, parsed_rows: list[ParsedRow], existing_ids: set[str]
    ) -> list[ImportRowResult]:
        """Validate all rows and return results."""
        results = []
        seen_ids: set[str] = set()

        parent_rows_by_email: dict[str, ParsedRow] = {}
        guest_rows_by_parent_email: dict[str, list[ParsedRow]] = {}
        guest_parent_emails: set[str] = set()
        guest_email_counts: dict[str, Counter[str]] = {}

        for row in parsed_rows:
            guest_of_email = self._normalize_email(row.data.get("guest_of_email"))
            if guest_of_email:
                guest_rows_by_parent_email.setdefault(guest_of_email, []).append(row)
                guest_parent_emails.add(guest_of_email)
                guest_email = self._normalize_email(row.data.get("registrant_email"))
                if guest_email:
                    guest_email_counts.setdefault(guest_of_email, Counter())[guest_email] += 1
            else:
                parent_email = self._normalize_email(row.data.get("registrant_email"))
                if parent_email:
                    parent_rows_by_email[parent_email] = row

        existing_parent_by_email = await self._fetch_registrations_by_email(
            event_id, guest_parent_emails
        )
        existing_guest_emails_by_registration = await self._fetch_guest_emails_by_registration(
            [registration.id for registration in existing_parent_by_email.values()]
        )
        food_option_name_index, food_option_id_index = await self._fetch_food_option_index(event_id)

        parent_capacity_by_email: dict[str, int] = {}
        for email, parent_row in parent_rows_by_email.items():
            parent_capacity_by_email[email] = max(self._resolve_guest_count(parent_row.data) - 1, 0)
        for email, registration in existing_parent_by_email.items():
            if email not in parent_capacity_by_email:
                total_guests = registration.number_of_guests or 1
                parent_capacity_by_email[email] = max(total_guests - 1, 0)

        for row in parsed_rows:
            issues: list[ValidationIssue] = []
            data = row.data
            guest_of_email = self._normalize_email(data.get("guest_of_email"))
            is_guest_row = bool(guest_of_email)

            # Validate required fields
            for field in REQUIRED_HEADERS:
                if is_guest_row and field == "external_registration_id":
                    continue
                value = data.get(field)
                if value is None or str(value).strip() == "":
                    issues.append(
                        ValidationIssue(
                            row_number=row.row_number,
                            severity=ValidationIssueSeverity.ERROR,
                            field_name=field,
                            message=f"Required field '{field}' is missing or empty",
                        )
                    )

            # Check for duplicates within file (non-guest rows only)
            external_id = str(data.get("external_registration_id", "")).strip()
            if external_id and not is_guest_row:
                if external_id in seen_ids:
                    issues.append(
                        ValidationIssue(
                            row_number=row.row_number,
                            severity=ValidationIssueSeverity.ERROR,
                            field_name="external_registration_id",
                            message=f"Duplicate external_registration_id in file: {external_id}",
                        )
                    )
                else:
                    seen_ids.add(external_id)

                # Check against existing IDs (warning only, will be skipped)
                if external_id in existing_ids:
                    issues.append(
                        ValidationIssue(
                            row_number=row.row_number,
                            severity=ValidationIssueSeverity.WARNING,
                            field_name="external_registration_id",
                            message=f"Registration with external_id '{external_id}' already exists and will be skipped",
                        )
                    )

            if is_guest_row:
                if (
                    guest_of_email not in parent_rows_by_email
                    and guest_of_email not in existing_parent_by_email
                ):
                    issues.append(
                        ValidationIssue(
                            row_number=row.row_number,
                            severity=ValidationIssueSeverity.ERROR,
                            field_name="guest_of_email",
                            message=f"No parent registration found for guest_of_email: {guest_of_email}",
                        )
                    )

                guest_email = self._normalize_email(data.get("registrant_email"))
                if guest_email and guest_of_email:
                    if guest_email_counts.get(guest_of_email, Counter()).get(guest_email, 0) > 1:
                        issues.append(
                            ValidationIssue(
                                row_number=row.row_number,
                                severity=ValidationIssueSeverity.ERROR,
                                field_name="registrant_email",
                                message="Duplicate guest email for the same registration",
                            )
                        )

                if guest_of_email in existing_parent_by_email and guest_email:
                    existing_parent = existing_parent_by_email[guest_of_email]
                    existing_guest_emails = existing_guest_emails_by_registration.get(
                        existing_parent.id, set()
                    )
                    if guest_email in existing_guest_emails:
                        issues.append(
                            ValidationIssue(
                                row_number=row.row_number,
                                severity=ValidationIssueSeverity.ERROR,
                                field_name="registrant_email",
                                message="Guest email already exists for this registration",
                            )
                        )

                if guest_of_email in parent_capacity_by_email:
                    max_guest_rows = parent_capacity_by_email[guest_of_email]
                    if len(guest_rows_by_parent_email.get(guest_of_email, [])) > max_guest_rows:
                        issues.append(
                            ValidationIssue(
                                row_number=row.row_number,
                                severity=ValidationIssueSeverity.ERROR,
                                field_name="guest_count",
                                message="Guest rows exceed guest_count for the parent registration",
                            )
                        )

            # Validate numeric fields
            quantity = data.get("quantity")
            if quantity is not None:
                try:
                    qty_int = int(quantity)
                    if qty_int < 1:
                        issues.append(
                            ValidationIssue(
                                row_number=row.row_number,
                                severity=ValidationIssueSeverity.ERROR,
                                field_name="quantity",
                                message="Quantity must be at least 1",
                            )
                        )
                except (ValueError, TypeError):
                    issues.append(
                        ValidationIssue(
                            row_number=row.row_number,
                            severity=ValidationIssueSeverity.ERROR,
                            field_name="quantity",
                            message=f"Invalid quantity value: {quantity}",
                        )
                    )

            food_option_raw = str(data.get("food_option", "")).strip()
            if food_option_raw:
                food_option_id = self._resolve_food_option_from_indexes(
                    food_option_raw, food_option_name_index, food_option_id_index
                )
                if not food_option_id:
                    issues.append(
                        ValidationIssue(
                            row_number=row.row_number,
                            severity=ValidationIssueSeverity.ERROR,
                            field_name="food_option",
                            message=f"Food option not found for event: {food_option_raw}",
                        )
                    )

            # Validate registration_date
            reg_date = data.get("registration_date")
            if reg_date is not None:
                try:
                    if isinstance(reg_date, str):
                        datetime.strptime(reg_date, "%Y-%m-%d")
                    elif not isinstance(reg_date, date):
                        raise ValueError("Invalid date type")
                except (ValueError, TypeError):
                    issues.append(
                        ValidationIssue(
                            row_number=row.row_number,
                            severity=ValidationIssueSeverity.ERROR,
                            field_name="registration_date",
                            message=f"Invalid registration_date (use YYYY-MM-DD): {reg_date}",
                        )
                    )

            if not is_guest_row:
                await self._validate_ticket_purchase_link(
                    event_id=event_id,
                    row_number=row.row_number,
                    data=data,
                    issues=issues,
                )

            # Check event_id mismatch (warning only)
            file_event_id = data.get("event_id")
            if file_event_id and str(file_event_id).strip():
                issues.append(
                    ValidationIssue(
                        row_number=row.row_number,
                        severity=ValidationIssueSeverity.WARNING,
                        field_name="event_id",
                        message=f"event_id in file ({file_event_id}) is ignored; importing to selected event",
                    )
                )

            # Determine status
            has_errors = any(i.severity == ValidationIssueSeverity.ERROR for i in issues)

            if has_errors:
                status = ImportRowStatus.ERROR
                message = f"Validation failed with {len([i for i in issues if i.severity == ValidationIssueSeverity.ERROR])} error(s)"
            elif external_id in existing_ids:
                status = ImportRowStatus.SKIPPED
                message = f"Skipped: Registration with external_id '{external_id}' already exists"
            else:
                status = ImportRowStatus.CREATED
                message = "Valid, will be created"

            results.append(
                ImportRowResult(
                    row_number=row.row_number,
                    external_id=external_id or None,
                    registrant_name=str(data.get("registrant_name", "")).strip() or None,
                    registrant_email=str(data.get("registrant_email", "")).strip() or None,
                    status=status,
                    message=message,
                    issues=issues,
                )
            )

        return results

    async def _create_registration(
        self,
        event_id: UUID,
        row: dict[str, Any],
        user_id: UUID,
        guest_rows: list[dict[str, Any]] | None = None,
        food_option_name_index: dict[str, UUID] | None = None,
        food_option_id_index: dict[str, UUID] | None = None,
    ) -> EventRegistration | None:
        """Create a registration record from validated row data."""
        registrant_name_raw = str(row.get("registrant_name", "")).strip()
        registrant_email_raw = str(row.get("registrant_email", "")).strip().lower()
        registrant_phone = str(row.get("registrant_phone", "")).strip() or None

        if not registrant_name_raw or not registrant_email_raw:
            raise RegistrationImportError("Registrant name and email are required")

        registrant_first_name, registrant_last_name = self._split_name(registrant_name_raw)

        user = await self._get_or_create_user(
            email=registrant_email_raw,
            first_name=registrant_first_name,
            last_name=registrant_last_name,
            phone=registrant_phone,
        )

        existing_result = await self.db.execute(
            select(EventRegistration).where(
                EventRegistration.event_id == event_id,
                EventRegistration.user_id == user.id,
            )
        )
        existing_registration = existing_result.scalar_one_or_none()
        if existing_registration:
            return None

        number_of_guests = self._resolve_guest_count(row)
        ticket_purchase_id = await self._resolve_ticket_purchase_id(event_id, row)

        registration = EventRegistration(
            user_id=user.id,
            event_id=event_id,
            number_of_guests=number_of_guests,
            ticket_purchase_id=ticket_purchase_id,
        )
        self.db.add(registration)
        await self.db.flush()

        primary_guest = RegistrationGuest(
            registration_id=registration.id,
            user_id=user.id,
            name=registrant_name_raw,
            email=registrant_email_raw,
            phone=registrant_phone,
            bidder_number=self._parse_optional_int(row.get("bidder_number")),
            table_number=self._parse_optional_int(row.get("table_number")),
            status=RegistrationStatus.CONFIRMED.value,
            is_primary=True,
        )
        self.db.add(primary_guest)
        await self.db.flush()

        registrant_food_option_id = self._resolve_food_option_from_indexes(
            str(row.get("food_option", "")).strip(),
            food_option_name_index or {},
            food_option_id_index or {},
        )
        self._add_meal_selection(registration.id, primary_guest.id, registrant_food_option_id)

        explicit_guest_rows = guest_rows or []
        for guest_row in explicit_guest_rows:
            guest_name = str(guest_row.get("registrant_name", "")).strip() or None
            guest_email = str(guest_row.get("registrant_email", "")).strip().lower() or None
            guest_phone = str(guest_row.get("registrant_phone", "")).strip() or None
            guest = RegistrationGuest(
                registration_id=registration.id,
                name=guest_name,
                email=guest_email,
                phone=guest_phone,
                bidder_number=self._parse_optional_int(guest_row.get("bidder_number")),
                table_number=self._parse_optional_int(guest_row.get("table_number")),
            )
            self.db.add(guest)
            await self.db.flush()
            guest_food_option_id = self._resolve_food_option_from_indexes(
                str(guest_row.get("food_option", "")).strip(),
                food_option_name_index or {},
                food_option_id_index or {},
            )
            self._add_meal_selection(registration.id, guest.id, guest_food_option_id)

        additional_guest_count = max(number_of_guests - 1 - len(explicit_guest_rows), 0)
        for _ in range(additional_guest_count):
            self.db.add(
                RegistrationGuest(
                    registration_id=registration.id,
                    name=None,
                    email=None,
                    phone=None,
                )
            )

        return registration

    async def _get_or_create_user(
        self,
        email: str,
        first_name: str,
        last_name: str,
        phone: str | None,
    ) -> User:
        result = await self.db.execute(select(User).where(User.email == email))
        existing = result.scalar_one_or_none()
        if existing:
            return existing

        donor_role_stmt = select(Base.metadata.tables["roles"].c.id).where(
            Base.metadata.tables["roles"].c.name == "donor"
        )
        role_result = await self.db.execute(donor_role_stmt)
        donor_role_id = role_result.scalar_one()

        user = User(
            email=email,
            first_name=first_name,
            last_name=last_name,
            phone=phone,
            email_verified=False,
            is_active=False,
            role_id=donor_role_id,
        )
        user.password_hash = hash_password(uuid.uuid4().hex)
        self.db.add(user)
        await self.db.flush()
        return user

    async def _create_guest_registration(
        self,
        registration_id: UUID,
        row: dict[str, Any],
        food_option_name_index: dict[str, UUID],
        food_option_id_index: dict[str, UUID],
    ) -> None:
        guest_name = str(row.get("registrant_name", "")).strip() or None
        guest_email = str(row.get("registrant_email", "")).strip().lower() or None
        guest_phone = str(row.get("registrant_phone", "")).strip() or None
        guest = RegistrationGuest(
            registration_id=registration_id,
            name=guest_name,
            email=guest_email,
            phone=guest_phone,
            bidder_number=self._parse_optional_int(row.get("bidder_number")),
            table_number=self._parse_optional_int(row.get("table_number")),
        )
        self.db.add(guest)
        await self.db.flush()
        guest_food_option_id = self._resolve_food_option_from_indexes(
            str(row.get("food_option", "")).strip(),
            food_option_name_index,
            food_option_id_index,
        )
        self._add_meal_selection(registration_id, guest.id, guest_food_option_id)

    async def _resolve_ticket_purchase_id(self, event_id: UUID, row: dict[str, Any]) -> UUID | None:
        raw_purchase_id = str(row.get("ticket_purchase_id", "")).strip()
        purchaser_email = str(row.get("ticket_purchaser_email", "")).strip()
        purchase_date_raw = row.get("ticket_purchase_date")

        if raw_purchase_id:
            # Try UUID first
            try:
                purchase_id = UUID(raw_purchase_id)
                purchase_result = await self.db.execute(
                    select(TicketPurchase).where(TicketPurchase.id == purchase_id)
                )
                purchase = purchase_result.scalar_one_or_none()
                if purchase and purchase.event_id == event_id:
                    return purchase.id
            except (ValueError, TypeError):
                # Not a UUID, try external_sale_id
                purchase_result = await self.db.execute(
                    select(TicketPurchase).where(
                        TicketPurchase.event_id == event_id,
                        TicketPurchase.external_sale_id == raw_purchase_id,
                    )
                )
                purchase = purchase_result.scalar_one_or_none()
                if not purchase:
                    raise RegistrationImportError("Ticket purchase not found")
                return purchase.id

        if purchaser_email or purchase_date_raw not in (None, ""):
            if not purchaser_email or purchase_date_raw in (None, ""):
                raise RegistrationImportError(
                    "Both ticket_purchaser_email and ticket_purchase_date are required"
                )

            purchase_date = self._parse_purchase_date(purchase_date_raw)
            if not purchase_date:
                raise RegistrationImportError(f"Invalid ticket_purchase_date: {purchase_date_raw}")

            purchase_query = (
                select(TicketPurchase)
                .where(TicketPurchase.event_id == event_id)
                .where(func.lower(TicketPurchase.purchaser_email) == purchaser_email.lower())
                .where(func.date(TicketPurchase.purchased_at) == purchase_date)
            )
            purchase_result = await self.db.execute(purchase_query)
            purchases = list(purchase_result.scalars().all())
            if len(purchases) == 0:
                raise RegistrationImportError("No matching ticket purchase found for email/date")
            if len(purchases) > 1:
                raise RegistrationImportError("Multiple ticket purchases found for email/date")
            return purchases[0].id

        return None

    def _split_name(self, full_name: str) -> tuple[str, str]:
        parts = [part for part in full_name.strip().split(" ") if part]
        if len(parts) == 0:
            return "Donor", "User"
        if len(parts) == 1:
            return parts[0], ""
        return parts[0], " ".join(parts[1:])

    def _resolve_guest_count(self, row: dict[str, Any]) -> int:
        guest_count = self._parse_optional_int(row.get("guest_count"))
        if guest_count and guest_count >= 1:
            return guest_count

        quantity = self._parse_optional_int(row.get("quantity"))
        if quantity and quantity >= 1:
            return quantity

        return 1

    def _parse_optional_int(self, value: Any) -> int | None:
        if value in (None, ""):
            return None
        try:
            return int(value)
        except (TypeError, ValueError):
            return None

    def _add_meal_selection(
        self,
        registration_id: UUID,
        guest_id: UUID | None,
        food_option_id: UUID | None,
    ) -> None:
        if not food_option_id:
            return

        self.db.add(
            MealSelection(
                registration_id=registration_id,
                guest_id=guest_id,
                food_option_id=food_option_id,
            )
        )

    async def _fetch_food_option_index(
        self, event_id: UUID
    ) -> tuple[dict[str, UUID], dict[str, UUID]]:
        result = await self.db.execute(select(FoodOption).where(FoodOption.event_id == event_id))
        options = result.scalars().all()
        name_index = {option.name.lower(): option.id for option in options if option.name}
        id_index = {str(option.id): option.id for option in options}
        return name_index, id_index

    def _resolve_food_option_from_indexes(
        self,
        raw_value: str,
        name_index: dict[str, UUID],
        id_index: dict[str, UUID],
    ) -> UUID | None:
        if not raw_value:
            return None

        if raw_value in id_index:
            return id_index[raw_value]

        return name_index.get(raw_value.lower())

    def _normalize_email(self, value: Any) -> str:
        return str(value or "").strip().lower()

    async def _fetch_registrations_by_email(
        self, event_id: UUID, emails: set[str]
    ) -> dict[str, EventRegistration]:
        if not emails:
            return {}

        normalized = {email.strip().lower() for email in emails if email}
        if not normalized:
            return {}

        result = await self.db.execute(
            select(EventRegistration, User.email)
            .join(User, User.id == EventRegistration.user_id)
            .where(EventRegistration.event_id == event_id)
            .where(func.lower(User.email).in_(normalized))
        )

        return {row[1].lower(): row[0] for row in result.all() if row[1]}

    async def _fetch_guest_emails_by_registration(
        self, registration_ids: list[UUID]
    ) -> dict[UUID, set[str]]:
        if not registration_ids:
            return {}

        result = await self.db.execute(
            select(RegistrationGuest.registration_id, RegistrationGuest.email)
            .where(RegistrationGuest.registration_id.in_(registration_ids))
            .where(RegistrationGuest.email.isnot(None))
        )

        emails_by_registration: dict[UUID, set[str]] = {}
        for registration_id, email in result.all():
            if not email:
                continue
            emails_by_registration.setdefault(registration_id, set()).add(email.lower())

        return emails_by_registration

    async def _validate_ticket_purchase_link(
        self,
        event_id: UUID,
        row_number: int,
        data: dict[str, Any],
        issues: list[ValidationIssue],
    ) -> None:
        """Validate optional ticket purchase link fields."""
        raw_purchase_id = str(data.get("ticket_purchase_id", "")).strip()
        purchaser_email = str(data.get("ticket_purchaser_email", "")).strip()
        purchase_date_raw = data.get("ticket_purchase_date")

        if raw_purchase_id:
            # Try UUID first
            try:
                purchase_id = UUID(raw_purchase_id)
                purchase_result = await self.db.execute(
                    select(TicketPurchase).where(TicketPurchase.id == purchase_id)
                )
                purchase = purchase_result.scalar_one_or_none()
                if not purchase:
                    issues.append(
                        ValidationIssue(
                            row_number=row_number,
                            severity=ValidationIssueSeverity.ERROR,
                            field_name="ticket_purchase_id",
                            message="Ticket purchase not found",
                        )
                    )
                    return
                if purchase.event_id != event_id:
                    issues.append(
                        ValidationIssue(
                            row_number=row_number,
                            severity=ValidationIssueSeverity.ERROR,
                            field_name="ticket_purchase_id",
                            message="Ticket purchase does not match event",
                        )
                    )
                return
            except (ValueError, TypeError):
                # Not a UUID, try external_sale_id
                purchase_result = await self.db.execute(
                    select(TicketPurchase).where(
                        TicketPurchase.event_id == event_id,
                        TicketPurchase.external_sale_id == raw_purchase_id,
                    )
                )
                purchase = purchase_result.scalar_one_or_none()
                if not purchase:
                    issues.append(
                        ValidationIssue(
                            row_number=row_number,
                            severity=ValidationIssueSeverity.ERROR,
                            field_name="ticket_purchase_id",
                            message=f"Ticket purchase not found for external_sale_id: {raw_purchase_id}",
                        )
                    )
                    return
                # Success, no error to add
                return

        if purchaser_email or purchase_date_raw is not None:
            if not purchaser_email or purchase_date_raw in (None, ""):
                issues.append(
                    ValidationIssue(
                        row_number=row_number,
                        severity=ValidationIssueSeverity.ERROR,
                        field_name="ticket_purchaser_email",
                        message="Both ticket_purchaser_email and ticket_purchase_date are required",
                    )
                )
                return

            purchase_date = self._parse_purchase_date(purchase_date_raw)
            if not purchase_date:
                issues.append(
                    ValidationIssue(
                        row_number=row_number,
                        severity=ValidationIssueSeverity.ERROR,
                        field_name="ticket_purchase_date",
                        message=f"Invalid ticket_purchase_date (use YYYY-MM-DD): {purchase_date_raw}",
                    )
                )
                return

            purchase_query = (
                select(TicketPurchase)
                .where(TicketPurchase.event_id == event_id)
                .where(func.lower(TicketPurchase.purchaser_email) == purchaser_email.lower())
                .where(func.date(TicketPurchase.purchased_at) == purchase_date)
            )
            purchase_result = await self.db.execute(purchase_query)
            purchases = list(purchase_result.scalars().all())

            if len(purchases) == 0:
                issues.append(
                    ValidationIssue(
                        row_number=row_number,
                        severity=ValidationIssueSeverity.ERROR,
                        field_name="ticket_purchaser_email",
                        message="No matching ticket purchase found for email/date",
                    )
                )
                return

            if len(purchases) > 1:
                issues.append(
                    ValidationIssue(
                        row_number=row_number,
                        severity=ValidationIssueSeverity.ERROR,
                        field_name="ticket_purchaser_email",
                        message="Multiple ticket purchases found for email/date",
                    )
                )

    def _parse_purchase_date(self, value: Any) -> date | None:
        """Parse purchase date from import value."""
        if isinstance(value, date) and not isinstance(value, datetime):
            return value
        if isinstance(value, datetime):
            return value.date()
        if isinstance(value, str):
            raw = value.strip()
            if not raw:
                return None
            try:
                return datetime.fromisoformat(raw).date()
            except ValueError:
                try:
                    return datetime.strptime(raw, "%Y-%m-%d").date()
                except ValueError:
                    return None
        return None

    def _build_report(self, results: list[ImportRowResult], file_type: str) -> ImportReport:
        """Build import report from results."""
        counter = Counter(result.status for result in results)

        return ImportReport(
            total_rows=len(results),
            valid_rows=counter[ImportRowStatus.CREATED],
            error_rows=sum(1 for r in results if r.status == ImportRowStatus.ERROR),
            warning_rows=sum(
                1
                for r in results
                if any(i.severity == ValidationIssueSeverity.WARNING for i in r.issues)
            ),
            created_count=counter[ImportRowStatus.CREATED],
            skipped_count=counter[ImportRowStatus.SKIPPED],
            failed_count=counter[ImportRowStatus.ERROR],
            rows=results,
            error_report_url=self._build_error_report_url(results),
            file_type=file_type,
        )

    def _build_error_report_url(self, results: list[ImportRowResult]) -> str | None:
        error_rows = [row for row in results if row.status == ImportRowStatus.ERROR]
        if not error_rows:
            return None

        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(
            [
                "row_number",
                "external_id",
                "registrant_name",
                "registrant_email",
                "status",
                "message",
                "issues",
            ]
        )
        for row in error_rows:
            issues_text = "; ".join(
                [
                    f"{issue.field_name or ''}{':' if issue.field_name else ''}{issue.message}"
                    for issue in row.issues
                ]
            )
            writer.writerow(
                [
                    row.row_number,
                    row.external_id or "",
                    row.registrant_name or "",
                    row.registrant_email or "",
                    row.status,
                    row.message,
                    issues_text,
                ]
            )

        csv_data = output.getvalue().encode("utf-8")
        encoded = base64.b64encode(csv_data).decode("utf-8")
        return f"data:text/csv;base64,{encoded}"

    def build_error_report(self, request: ImportErrorReportRequest) -> ImportErrorReportResponse:
        """Build downloadable error report content."""
        rows = [row for row in request.rows if row.status == ImportRowStatus.ERROR]

        if request.format == ImportErrorReportFormat.JSON:
            payload = [row.model_dump() for row in rows]
            content = json.dumps(payload, indent=2)
            return ImportErrorReportResponse(
                format=request.format,
                content_type="application/json",
                filename="registration-import-errors.json",
                content=content,
            )

        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(
            [
                "row_number",
                "external_id",
                "registrant_name",
                "registrant_email",
                "status",
                "message",
                "issues",
            ]
        )
        for row in rows:
            issues_text = "; ".join(
                [
                    f"{issue.field_name or ''}{':' if issue.field_name else ''}{issue.message}"
                    for issue in row.issues
                ]
            )
            writer.writerow(
                [
                    row.row_number,
                    row.external_id or "",
                    row.registrant_name or "",
                    row.registrant_email or "",
                    row.status,
                    row.message,
                    issues_text,
                ]
            )

        return ImportErrorReportResponse(
            format=request.format,
            content_type="text/csv",
            filename="registration-import-errors.csv",
            content=output.getvalue(),
        )
