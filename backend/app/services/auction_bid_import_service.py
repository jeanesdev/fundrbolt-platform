"""Service for auction bid bulk import."""

from __future__ import annotations

import csv
import hashlib
import io
import json
import time
from collections.abc import Sequence
from dataclasses import dataclass
from datetime import UTC, datetime
from decimal import Decimal, InvalidOperation
from typing import Any
from uuid import UUID

import magic
from openpyxl import load_workbook
from sqlalchemy import func, select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.sql.selectable import Subquery

from app.core.logging import get_logger
from app.models.auction_bid import AuctionBid, BidStatus, BidType, TransactionStatus
from app.models.auction_bid_import import (
    AuctionBidImportBatch,
    AuctionBidImportFormat,
    AuctionBidImportIssue,
    AuctionBidImportIssueSeverity,
    AuctionBidImportStatus,
)
from app.models.auction_item import AuctionItem
from app.models.event_registration import EventRegistration
from app.models.registration_guest import RegistrationGuest
from app.models.user import User
from app.schemas.auction_bid_import import (
    AuctionBidDashboardHighestBid,
    AuctionBidDashboardRecentBid,
    AuctionBidDashboardResponse,
    AuctionBidImportSummary,
    AuctionBidPreflightResult,
)
from app.schemas.auction_bid_import import (
    AuctionBidImportIssue as SchemaIssue,
)
from app.schemas.auction_bid_import import (
    AuctionBidImportIssueSeverity as SchemaIssueSeverity,
)

MAX_IMPORT_ROWS = 10000
REQUIRED_HEADERS = ["donor_email", "auction_item_code", "bid_amount", "bid_time"]
logger = get_logger(__name__)


@dataclass
class ParsedRow:
    row_number: int
    data: dict[str, Any]


@dataclass
class DonorLookup:
    users_by_email: dict[str, User]
    guest_emails_without_user: set[str]


class AuctionBidImportError(Exception):
    """Base exception for auction bid import errors."""


class AuctionBidImportService:
    """Bulk import service for auction bids."""

    def __init__(self, db: AsyncSession) -> None:
        self.db = db

    async def preflight(
        self,
        event_id: UUID,
        file_bytes: bytes,
        filename: str,
        created_by: UUID,
    ) -> AuctionBidPreflightResult:
        start_time = time.monotonic()
        logger.info(
            "Auction bid preflight started",
            extra={
                "event_id": str(event_id),
                "file_name": filename,
                "file_size_bytes": len(file_bytes),
            },
        )

        detect_start = time.monotonic()
        detected_format = self._detect_format(file_bytes, filename)
        parsed_rows = self._parse_file(file_bytes, detected_format)
        logger.info(
            "Auction bid preflight parsed",
            extra={
                "event_id": str(event_id),
                "format": detected_format.value,
                "row_count": len(parsed_rows),
                "elapsed_ms": int((time.monotonic() - detect_start) * 1000),
            },
        )

        if len(parsed_rows) > MAX_IMPORT_ROWS:
            raise AuctionBidImportError(
                f"File contains {len(parsed_rows)} rows. Maximum allowed is {MAX_IMPORT_ROWS}."
            )

        if not parsed_rows:
            raise AuctionBidImportError("File contains no data rows.")

        donor_emails: set[str] = {
            email
            for row in parsed_rows
            if (email := self._normalize_email(row.data.get("donor_email")))
        }
        item_codes: set[str] = {
            code
            for row in parsed_rows
            if (code := self._normalize_item_code(row.data.get("auction_item_code")))
        }

        lookup_start = time.monotonic()
        donor_lookup = await self._fetch_donors_by_email(event_id, donor_emails)
        item_lookup = await self._fetch_items_by_code(event_id, item_codes)
        bidder_lookup = await self._fetch_bidder_numbers(event_id, donor_lookup.users_by_email)
        current_high_bids = await self._fetch_current_high_bids(item_lookup)
        logger.info(
            "Auction bid preflight lookups",
            extra={
                "event_id": str(event_id),
                "donor_count": len(donor_lookup.users_by_email),
                "guest_missing_user_count": len(donor_lookup.guest_emails_without_user),
                "item_count": len(item_lookup),
                "bidder_count": len(bidder_lookup),
                "elapsed_ms": int((time.monotonic() - lookup_start) * 1000),
            },
        )

        duplicate_counts = self._build_duplicate_counts(parsed_rows)
        existing_fingerprints = await self._fetch_existing_bid_fingerprints(
            event_id, donor_lookup.users_by_email, item_lookup
        )

        issues: list[SchemaIssue] = []
        warnings: list[SchemaIssue] = []
        valid_count = 0
        error_count = 0
        warning_count = 0

        updated_high_bids = dict(current_high_bids)
        for parsed_row in parsed_rows:
            row_issues = self._validate_row(
                parsed_row,
                donor_lookup.users_by_email,
                donor_lookup.guest_emails_without_user,
                item_lookup,
                bidder_lookup,
                updated_high_bids,
                duplicate_counts,
                existing_fingerprints,
            )
            if row_issues:
                has_error = any(issue.severity == SchemaIssueSeverity.ERROR for issue in row_issues)
                if has_error:
                    error_count += 1
                if not has_error:
                    warning_count += 1
                for issue in row_issues:
                    if issue.severity == SchemaIssueSeverity.ERROR:
                        issues.append(issue)
                    else:
                        warnings.append(issue)
            else:
                valid_count += 1

            if not row_issues:
                item_code = self._normalize_item_code(parsed_row.data.get("auction_item_code"))
                donor_email = self._normalize_email(parsed_row.data.get("donor_email"))
                bid_amount = self._parse_decimal(parsed_row.data.get("bid_amount"))
                bid_time = self._parse_bid_time(parsed_row.data.get("bid_time"))
                item = item_lookup.get(item_code or "")
                donor = donor_lookup.users_by_email.get(donor_email or "")
                if item and donor:
                    updated_high_bids[item.id] = AuctionBid(
                        event_id=item.event_id,
                        auction_item_id=item.id,
                        user_id=donor.id,
                        bidder_number=bidder_lookup.get(donor.id, 0),
                        bid_amount=bid_amount,
                        max_bid=None,
                        bid_type=BidType.REGULAR.value,
                        bid_status=BidStatus.ACTIVE.value,
                        transaction_status=TransactionStatus.PENDING.value,
                        placed_at=bid_time,
                        source_bid_id=None,
                        created_by=created_by,
                    )

        checksum = hashlib.sha256(file_bytes).hexdigest()

        batch = AuctionBidImportBatch(
            event_id=event_id,
            created_by=created_by,
            status=AuctionBidImportStatus.PREFLIGHTED,
            source_filename=filename,
            source_format=detected_format,
            row_count=len(parsed_rows),
            valid_count=valid_count,
            error_count=error_count,
            warning_count=warning_count,
            preflight_checksum=checksum,
        )
        self.db.add(batch)
        await self.db.flush()

        for issue in issues + warnings:
            issue_record = AuctionBidImportIssue(
                batch_id=batch.id,
                row_number=issue.row_number,
                field_name=issue.field_name,
                severity=AuctionBidImportIssueSeverity(issue.severity.value),
                message=issue.message,
                raw_value=issue.raw_value,
            )
            self.db.add(issue_record)

        result = AuctionBidPreflightResult(
            import_batch_id=str(batch.id),
            detected_format=detected_format.value,
            total_rows=len(parsed_rows),
            valid_rows=valid_count,
            invalid_rows=error_count,
            warning_rows=warning_count,
            row_errors=issues,
            row_warnings=warnings,
            error_report_url=None,
        )
        logger.info(
            "Auction bid preflight completed",
            extra={
                "event_id": str(event_id),
                "total_rows": result.total_rows,
                "invalid_rows": result.invalid_rows,
                "warning_rows": result.warning_rows,
                "elapsed_ms": int((time.monotonic() - start_time) * 1000),
            },
        )
        return result

    async def confirm_import(
        self,
        event_id: UUID,
        import_batch_id: UUID,
        file_bytes: bytes,
        user_id: UUID,
    ) -> AuctionBidImportSummary:
        result = await self.db.execute(
            select(AuctionBidImportBatch).where(
                AuctionBidImportBatch.id == import_batch_id,
                AuctionBidImportBatch.event_id == event_id,
            )
        )
        batch = result.scalar_one_or_none()
        if not batch:
            raise AuctionBidImportError("Preflight batch not found")

        checksum = hashlib.sha256(file_bytes).hexdigest()
        if batch.preflight_checksum != checksum:
            raise AuctionBidImportError(
                "File has changed since preflight. Please re-run preflight."
            )

        if batch.error_count > 0:
            raise AuctionBidImportError(
                "Cannot import file with errors. Please fix errors and re-run preflight."
            )

        parsed_rows = self._parse_file(file_bytes, batch.source_format)

        donor_emails: set[str] = {
            email
            for row in parsed_rows
            if (email := self._normalize_email(row.data.get("donor_email")))
        }
        item_codes: set[str] = {
            code
            for row in parsed_rows
            if (code := self._normalize_item_code(row.data.get("auction_item_code")))
        }

        donor_lookup = await self._fetch_donors_by_email(event_id, donor_emails)
        item_lookup = await self._fetch_items_by_code(event_id, item_codes)
        bidder_lookup = await self._fetch_bidder_numbers(event_id, donor_lookup.users_by_email)
        current_high_bids = await self._fetch_current_high_bids(item_lookup)
        duplicate_counts = self._build_duplicate_counts(parsed_rows)
        existing_fingerprints = await self._fetch_existing_bid_fingerprints(
            event_id, donor_lookup.users_by_email, item_lookup
        )

        created_count = 0
        skipped_count = 0
        started_at = datetime.now(UTC)
        updated_high_bids = dict(current_high_bids)

        for parsed_row in parsed_rows:
            row_issues = self._validate_row(
                parsed_row,
                donor_lookup.users_by_email,
                donor_lookup.guest_emails_without_user,
                item_lookup,
                bidder_lookup,
                updated_high_bids,
                duplicate_counts,
                existing_fingerprints,
            )
            if any(issue.severity == SchemaIssueSeverity.ERROR for issue in row_issues):
                raise AuctionBidImportError(
                    f"Preflight mismatch detected at row {parsed_row.row_number}."
                )

            if any(issue.severity == SchemaIssueSeverity.WARNING for issue in row_issues):
                if any(
                    issue.message == "Bid already exists and will be skipped"
                    for issue in row_issues
                ):
                    skipped_count += 1
                    continue

            donor_email = self._normalize_email(parsed_row.data.get("donor_email"))
            item_code = self._normalize_item_code(parsed_row.data.get("auction_item_code"))
            if not donor_email or not item_code:
                raise AuctionBidImportError(
                    f"Missing donor or item at row {parsed_row.row_number}."
                )

            donor = donor_lookup.users_by_email[donor_email]
            item = item_lookup[item_code]
            bidder_number = bidder_lookup.get(donor.id)
            if bidder_number is None:
                raise AuctionBidImportError(f"Bidder number not found for donor {donor_email}.")

            bid_amount = self._parse_decimal(parsed_row.data.get("bid_amount"))
            bid_time = self._parse_bid_time(parsed_row.data.get("bid_time"))

            new_bid = AuctionBid(
                event_id=event_id,
                auction_item_id=item.id,
                user_id=donor.id,
                bidder_number=bidder_number,
                bid_amount=bid_amount,
                max_bid=None,
                bid_type=BidType.REGULAR.value,
                bid_status=BidStatus.ACTIVE.value,
                transaction_status=TransactionStatus.PENDING.value,
                placed_at=bid_time,
                source_bid_id=None,
                created_by=user_id,
            )
            self.db.add(new_bid)
            await self.db.flush()

            current_high = updated_high_bids.get(item.id)
            if current_high and current_high.user_id != donor.id:
                outbid = AuctionBid(
                    event_id=current_high.event_id,
                    auction_item_id=current_high.auction_item_id,
                    user_id=current_high.user_id,
                    bidder_number=current_high.bidder_number,
                    bid_amount=current_high.bid_amount,
                    max_bid=current_high.max_bid,
                    bid_type=current_high.bid_type,
                    bid_status=BidStatus.OUTBID.value,
                    transaction_status=current_high.transaction_status,
                    placed_at=bid_time,
                    source_bid_id=current_high.id,
                    created_by=user_id,
                )
                self.db.add(outbid)

            updated_high_bids[item.id] = new_bid
            created_count += 1

        batch.status = AuctionBidImportStatus.IMPORTED
        batch.created_by = user_id
        batch.created_count = created_count
        batch.skipped_count = skipped_count
        batch.completed_at = datetime.now(UTC)
        await self.db.flush()

        completed_at = batch.completed_at or datetime.now(UTC)
        return AuctionBidImportSummary(
            import_batch_id=str(batch.id),
            created_bids=created_count,
            skipped_bids=skipped_count,
            started_at=started_at,
            completed_at=completed_at,
        )

    async def get_dashboard(self, event_id: UUID) -> AuctionBidDashboardResponse:
        total_bid_count = (
            await self.db.execute(
                select(func.count(AuctionBid.id)).where(AuctionBid.event_id == event_id)
            )
        ).scalar_one()
        total_bid_value = (
            await self.db.execute(
                select(func.coalesce(func.sum(AuctionBid.bid_amount), 0)).where(
                    AuctionBid.event_id == event_id
                )
            )
        ).scalar_one()

        latest_ids = await self._latest_bid_records_for_event_subquery(event_id)

        highest_query = (
            select(AuctionBid, AuctionItem, User.email)
            .join(AuctionItem, AuctionBid.auction_item_id == AuctionItem.id)
            .join(User, AuctionBid.user_id == User.id)
            .where(
                AuctionBid.id.in_(select(latest_ids.c.id)),
                AuctionBid.bid_status.in_([BidStatus.ACTIVE.value, BidStatus.WINNING.value]),
            )
        )
        highest_rows = (await self.db.execute(highest_query)).all()
        highest_map: dict[UUID, AuctionBidDashboardHighestBid] = {}
        for bid, item, email in highest_rows:
            existing = highest_map.get(item.id)
            if not existing or bid.bid_amount > existing.bid_amount:
                highest_map[item.id] = AuctionBidDashboardHighestBid(
                    auction_item_code=item.external_id,
                    bid_amount=bid.bid_amount,
                    bidder_email=email,
                )

        recent_query = (
            select(AuctionBid, AuctionItem, User.email)
            .join(AuctionItem, AuctionBid.auction_item_id == AuctionItem.id)
            .join(User, AuctionBid.user_id == User.id)
            .where(AuctionBid.event_id == event_id)
            .order_by(AuctionBid.placed_at.desc())
            .limit(10)
        )
        recent_rows = (await self.db.execute(recent_query)).all()
        recent_bids = [
            AuctionBidDashboardRecentBid(
                auction_item_code=item.external_id,
                bid_amount=bid.bid_amount,
                bidder_email=email,
                bid_time=bid.placed_at,
            )
            for bid, item, email in recent_rows
        ]

        return AuctionBidDashboardResponse(
            total_bid_count=total_bid_count or 0,
            total_bid_value=total_bid_value or Decimal("0"),
            highest_bids=list(highest_map.values()),
            recent_bids=recent_bids,
        )

    def _detect_format(self, file_bytes: bytes, filename: str) -> AuctionBidImportFormat:
        mime_type = magic.from_buffer(file_bytes, mime=True)

        if mime_type == "text/csv" or filename.lower().endswith(".csv"):
            return AuctionBidImportFormat.CSV
        if mime_type == "application/json" or filename.lower().endswith(".json"):
            return AuctionBidImportFormat.JSON
        if mime_type in [
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "application/vnd.ms-excel",
        ] or filename.lower().endswith((".xlsx", ".xls")):
            return AuctionBidImportFormat.XLSX
        raise AuctionBidImportError(
            "Unsupported file format. Please upload a CSV, JSON, or Excel file."
        )

    def _parse_file(
        self, file_bytes: bytes, file_format: AuctionBidImportFormat
    ) -> list[ParsedRow]:
        if file_format == AuctionBidImportFormat.CSV:
            return self._parse_csv(file_bytes)
        if file_format == AuctionBidImportFormat.JSON:
            return self._parse_json(file_bytes)
        if file_format == AuctionBidImportFormat.XLSX:
            return self._parse_xlsx(file_bytes)
        raise AuctionBidImportError(f"Unsupported format: {file_format}")

    def _parse_csv(self, file_bytes: bytes) -> list[ParsedRow]:
        try:
            text = file_bytes.decode("utf-8")
            reader = csv.DictReader(io.StringIO(text))
            self._validate_required_headers(reader.fieldnames or [])
            rows: list[ParsedRow] = []
            for idx, row in enumerate(reader, start=2):
                if self._is_empty_row(row):
                    continue
                rows.append(ParsedRow(row_number=idx, data=dict(row)))
            return rows
        except UnicodeDecodeError as exc:
            raise AuctionBidImportError("CSV must be UTF-8 encoded") from exc
        except Exception as exc:
            raise AuctionBidImportError(f"Failed to parse CSV: {exc}") from exc

    def _parse_json(self, file_bytes: bytes) -> list[ParsedRow]:
        try:
            text = file_bytes.decode("utf-8")
            data = json.loads(text)
            if not isinstance(data, list):
                raise AuctionBidImportError("JSON must contain an array of bids")
            rows = []
            for idx, item in enumerate(data, start=1):
                if not isinstance(item, dict):
                    raise AuctionBidImportError("JSON rows must be objects")
                rows.append(ParsedRow(row_number=idx, data=item))
            return rows
        except json.JSONDecodeError as exc:
            raise AuctionBidImportError(f"Invalid JSON: {exc}") from exc
        except Exception as exc:
            raise AuctionBidImportError(f"Failed to parse JSON: {exc}") from exc

    def _parse_xlsx(self, file_bytes: bytes) -> list[ParsedRow]:
        try:
            workbook = load_workbook(io.BytesIO(file_bytes))
            sheet = workbook.active
            if sheet is None:
                raise AuctionBidImportError("Excel file has no active sheet")
            header_row = [cell.value for cell in sheet[1]]
            headers = [self._normalize_header(value) for value in header_row]
            self._validate_required_headers(headers)

            rows: list[ParsedRow] = []
            for idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                if self._is_empty_row(row):
                    continue
                data = {headers[i]: row[i] for i in range(len(headers)) if i < len(row)}
                rows.append(ParsedRow(row_number=idx, data=data))
            return rows
        except Exception as exc:
            raise AuctionBidImportError(f"Failed to parse Excel: {exc}") from exc

    def _validate_required_headers(self, headers: Sequence[Any]) -> None:
        normalized = {self._normalize_header(header) for header in headers if header}
        missing = [field for field in REQUIRED_HEADERS if field not in normalized]
        if missing:
            raise AuctionBidImportError(f"Missing required columns: {', '.join(missing)}")

    def _validate_row(
        self,
        parsed_row: ParsedRow,
        donor_lookup: dict[str, User],
        guest_emails_without_user: set[str],
        item_lookup: dict[str, AuctionItem],
        bidder_lookup: dict[UUID, int],
        current_high_bids: dict[UUID, AuctionBid | None],
        duplicate_counts: dict[tuple[str, str, str, str], int],
        existing_fingerprints: set[tuple[str, str, str, str]],
    ) -> list[SchemaIssue]:
        issues: list[SchemaIssue] = []
        data = parsed_row.data
        row_num = parsed_row.row_number

        for field in REQUIRED_HEADERS:
            value = data.get(field)
            if value is None or str(value).strip() == "":
                issues.append(
                    SchemaIssue(
                        row_number=row_num,
                        field_name=field,
                        severity=SchemaIssueSeverity.ERROR,
                        message=f"Missing required field: {field}",
                    )
                )

        donor_email = self._normalize_email(data.get("donor_email"))
        if donor_email and donor_email not in donor_lookup:
            if donor_email in guest_emails_without_user:
                issues.append(
                    SchemaIssue(
                        row_number=row_num,
                        field_name="donor_email",
                        severity=SchemaIssueSeverity.ERROR,
                        message=(
                            f"Donor email '{donor_email}' is linked to a guest without a user account"
                        ),
                        raw_value=donor_email,
                    )
                )
            else:
                issues.append(
                    SchemaIssue(
                        row_number=row_num,
                        field_name="donor_email",
                        severity=SchemaIssueSeverity.ERROR,
                        message=f"Donor email '{donor_email}' not found for this event",
                        raw_value=donor_email,
                    )
                )

        item_code = self._normalize_item_code(data.get("auction_item_code"))
        if item_code and item_code not in item_lookup:
            issues.append(
                SchemaIssue(
                    row_number=row_num,
                    field_name="auction_item_code",
                    severity=SchemaIssueSeverity.ERROR,
                    message=f"Auction item code '{item_code}' not found",
                    raw_value=item_code,
                )
            )

        if donor_email and donor_email in donor_lookup:
            bidder_number = bidder_lookup.get(donor_lookup[donor_email].id)
            if bidder_number is None:
                issues.append(
                    SchemaIssue(
                        row_number=row_num,
                        field_name="donor_email",
                        severity=SchemaIssueSeverity.ERROR,
                        message=f"Bidder number not found for donor '{donor_email}'",
                        raw_value=donor_email,
                    )
                )

        bid_amount = None
        if data.get("bid_amount") not in (None, ""):
            try:
                bid_amount = self._parse_decimal(data.get("bid_amount"))
                if bid_amount < 0:
                    raise InvalidOperation
            except (InvalidOperation, ValueError, TypeError):
                issues.append(
                    SchemaIssue(
                        row_number=row_num,
                        field_name="bid_amount",
                        severity=SchemaIssueSeverity.ERROR,
                        message="Bid amount must be a valid non-negative number",
                        raw_value=str(data.get("bid_amount")),
                    )
                )

        bid_time = None
        if data.get("bid_time") not in (None, ""):
            try:
                bid_time = self._parse_bid_time(data.get("bid_time"))
            except ValueError as exc:
                issues.append(
                    SchemaIssue(
                        row_number=row_num,
                        field_name="bid_time",
                        severity=SchemaIssueSeverity.ERROR,
                        message=str(exc),
                        raw_value=str(data.get("bid_time")),
                    )
                )

        if donor_email and item_code:
            duplicate_key = (
                donor_email,
                item_code,
                str(data.get("bid_amount", "")).strip(),
                str(data.get("bid_time", "")).strip(),
            )
            if duplicate_counts.get(duplicate_key, 0) > 1:
                issues.append(
                    SchemaIssue(
                        row_number=row_num,
                        field_name=None,
                        severity=SchemaIssueSeverity.ERROR,
                        message="Duplicate bid row detected in file",
                    )
                )

        if donor_email and item_code and bid_amount is not None and bid_time is not None:
            donor = donor_lookup.get(donor_email)
            item = item_lookup.get(item_code)
            if donor and item:
                fingerprint = self._build_bid_fingerprint(donor.id, item.id, bid_amount, bid_time)
                if fingerprint in existing_fingerprints:
                    issues.append(
                        SchemaIssue(
                            row_number=row_num,
                            field_name=None,
                            severity=SchemaIssueSeverity.WARNING,
                            message="Bid already exists and will be skipped",
                        )
                    )

        if item_code and bid_amount is not None:
            item = item_lookup.get(item_code)
            if item:
                current_high = current_high_bids.get(item.id)
                min_bid = item.starting_bid
                if current_high and current_high.bid_amount is not None:
                    min_bid = current_high.bid_amount + item.bid_increment
                if bid_amount < min_bid:
                    issues.append(
                        SchemaIssue(
                            row_number=row_num,
                            field_name="bid_amount",
                            severity=SchemaIssueSeverity.ERROR,
                            message=(f"Bid amount must be at least {min_bid} for item {item_code}"),
                            raw_value=str(data.get("bid_amount")),
                        )
                    )

        return issues

    async def _fetch_donors_by_email(self, event_id: UUID, emails: set[str]) -> DonorLookup:
        if not emails:
            return DonorLookup(users_by_email={}, guest_emails_without_user=set())

        registrant_stmt = (
            select(User)
            .join(EventRegistration, EventRegistration.user_id == User.id)
            .where(EventRegistration.event_id == event_id, User.email.in_(emails))
        )
        registrant_result = await self.db.execute(registrant_stmt)
        registrant_users = registrant_result.scalars().all()
        users_by_email: dict[str, User] = {
            user.email.lower(): user for user in registrant_users if user.email
        }

        guest_stmt = (
            select(RegistrationGuest.email, User)
            .join(EventRegistration, RegistrationGuest.registration_id == EventRegistration.id)
            .outerjoin(User, RegistrationGuest.user_id == User.id)
            .where(EventRegistration.event_id == event_id, RegistrationGuest.email.in_(emails))
        )
        guest_result = await self.db.execute(guest_stmt)
        guest_emails_without_user: set[str] = set()
        for guest_email, user in guest_result.all():
            normalized_email = self._normalize_email(guest_email)
            if not normalized_email:
                continue
            if user is None:
                if normalized_email not in users_by_email:
                    guest_emails_without_user.add(normalized_email)
                continue
            users_by_email.setdefault(normalized_email, user)

        return DonorLookup(
            users_by_email=users_by_email,
            guest_emails_without_user=guest_emails_without_user,
        )

    async def _fetch_items_by_code(self, event_id: UUID, codes: set[str]) -> dict[str, AuctionItem]:
        if not codes:
            return {}
        result = await self.db.execute(
            select(AuctionItem).where(
                AuctionItem.event_id == event_id,
                AuctionItem.external_id.in_(codes),
            )
        )
        items = result.scalars().all()
        return {item.external_id.lower(): item for item in items}

    async def _fetch_bidder_numbers(
        self, event_id: UUID, donors: dict[str, User]
    ) -> dict[UUID, int]:
        if not donors:
            return {}
        user_ids = [user.id for user in donors.values()]
        primary_stmt = (
            select(RegistrationGuest.user_id, RegistrationGuest.bidder_number)
            .join(EventRegistration, RegistrationGuest.registration_id == EventRegistration.id)
            .where(
                EventRegistration.event_id == event_id,
                RegistrationGuest.user_id.in_(user_ids),
                RegistrationGuest.bidder_number.isnot(None),
                RegistrationGuest.is_primary.is_(True),
            )
        )
        primary_result = await self.db.execute(primary_stmt)
        bidder_numbers: dict[UUID, int] = {
            user_id: bidder_number
            for user_id, bidder_number in primary_result.all()
            if user_id and bidder_number is not None
        }

        missing_users = [user_id for user_id in user_ids if user_id not in bidder_numbers]
        if missing_users:
            fallback_stmt = (
                select(RegistrationGuest.user_id, RegistrationGuest.bidder_number)
                .join(EventRegistration, RegistrationGuest.registration_id == EventRegistration.id)
                .where(
                    EventRegistration.event_id == event_id,
                    RegistrationGuest.user_id.in_(missing_users),
                    RegistrationGuest.bidder_number.isnot(None),
                )
            )
            fallback_result = await self.db.execute(fallback_stmt)
            for user_id, bidder_number in fallback_result.all():
                if user_id and bidder_number is not None:
                    bidder_numbers.setdefault(user_id, bidder_number)

        return bidder_numbers

    async def _fetch_existing_bid_fingerprints(
        self,
        event_id: UUID,
        donors: dict[str, User],
        items: dict[str, AuctionItem],
    ) -> set[tuple[str, str, str, str]]:
        if not donors or not items:
            return set()
        user_ids = [user.id for user in donors.values()]
        item_ids = [item.id for item in items.values()]
        stmt = select(AuctionBid).where(
            AuctionBid.event_id == event_id,
            AuctionBid.user_id.in_(user_ids),
            AuctionBid.auction_item_id.in_(item_ids),
        )
        result = await self.db.execute(stmt)
        bids = result.scalars().all()
        return {
            self._build_bid_fingerprint(
                bid.user_id, bid.auction_item_id, bid.bid_amount, bid.placed_at
            )
            for bid in bids
        }

    async def _fetch_current_high_bids(
        self, items: dict[str, AuctionItem]
    ) -> dict[UUID, AuctionBid | None]:
        if not items:
            return {}
        item_ids = [item.id for item in items.values()]
        latest_subq = (
            select(AuctionBid.source_bid_id).where(AuctionBid.source_bid_id.isnot(None)).subquery()
        )
        query = select(AuctionBid).where(
            AuctionBid.auction_item_id.in_(item_ids),
            AuctionBid.id.not_in(select(latest_subq.c.source_bid_id)),
        )
        result = await self.db.execute(query)
        bids = result.scalars().all()
        current_high: dict[UUID, AuctionBid | None] = dict.fromkeys(item_ids)
        for bid in bids:
            if bid.bid_status not in {BidStatus.ACTIVE.value, BidStatus.WINNING.value}:
                continue
            existing = current_high.get(bid.auction_item_id)
            if not existing or bid.bid_amount > existing.bid_amount:
                current_high[bid.auction_item_id] = bid
        return current_high

    async def _latest_bid_records_for_event_subquery(self, event_id: UUID) -> Subquery:
        source_subq = (
            select(AuctionBid.source_bid_id)
            .where(
                AuctionBid.event_id == event_id,
                AuctionBid.source_bid_id.isnot(None),
            )
            .subquery()
        )
        return (
            select(AuctionBid.id)
            .where(
                AuctionBid.event_id == event_id,
                AuctionBid.id.not_in(select(source_subq.c.source_bid_id)),
            )
            .subquery()
        )

    def _build_duplicate_counts(
        self, parsed_rows: list[ParsedRow]
    ) -> dict[tuple[str, str, str, str], int]:
        counts: dict[tuple[str, str, str, str], int] = {}
        for row in parsed_rows:
            donor_email = self._normalize_email(row.data.get("donor_email")) or ""
            item_code = self._normalize_item_code(row.data.get("auction_item_code")) or ""
            amount = str(row.data.get("bid_amount", "")).strip()
            time_value = str(row.data.get("bid_time", "")).strip()
            key = (donor_email, item_code, amount, time_value)
            counts[key] = counts.get(key, 0) + 1
        return counts

    @staticmethod
    def _normalize_email(value: Any) -> str | None:
        if value is None:
            return None
        email = str(value).strip().lower()
        return email or None

    @staticmethod
    def _normalize_item_code(value: Any) -> str | None:
        if value is None:
            return None
        code = str(value).strip().lower()
        return code or None

    @staticmethod
    def _normalize_header(value: Any) -> str:
        if value is None:
            return ""
        return str(value).strip().lower()

    @staticmethod
    def _parse_decimal(value: Any) -> Decimal:
        if isinstance(value, Decimal):
            return value
        return Decimal(str(value))

    @staticmethod
    def _parse_bid_time(value: Any) -> datetime:
        if isinstance(value, datetime):
            return value
        bid_time_str = str(value).strip()
        if not bid_time_str:
            raise ValueError("Bid time is required")
        try:
            parsed = datetime.fromisoformat(bid_time_str)
            return parsed if parsed.tzinfo else parsed.replace(tzinfo=UTC)
        except ValueError:
            try:
                from dateutil import parser as date_parser

                parsed = date_parser.parse(bid_time_str)
                return parsed if parsed.tzinfo else parsed.replace(tzinfo=UTC)
            except Exception as exc:
                raise ValueError("Bid time must be a valid date-time value") from exc

    @staticmethod
    def _build_bid_fingerprint(
        user_id: UUID, item_id: UUID, bid_amount: Decimal, bid_time: datetime
    ) -> tuple[str, str, str, str]:
        return (str(user_id), str(item_id), str(bid_amount), bid_time.isoformat())

    @staticmethod
    def _is_empty_row(row: Any) -> bool:
        if isinstance(row, dict):
            return all(str(value).strip() == "" for value in row.values())
        if isinstance(row, (list, tuple)):
            return all(value is None or str(value).strip() == "" for value in row)
        return False
