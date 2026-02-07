"""Service for auction item bulk import."""

from __future__ import annotations

import base64
import csv
import io
from collections import Counter
from dataclasses import dataclass
from decimal import Decimal
from typing import Any
from uuid import UUID

import magic
from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.auction_item_import import ALLOWED_CATEGORIES, MAX_IMPORT_ROWS
from app.core.config import Settings, get_settings
from app.models.auction_item import (
    AuctionItem,
    AuctionItemMedia,
    AuctionType,
    ItemStatus,
    MediaType,
)
from app.schemas.auction_item_import import (
    AuctionItemImportRow,
    ImportImageStatus,
    ImportReport,
    ImportRowResult,
    ImportRowStatus,
)
from app.services.auction_item_import_zip import ImportZipValidationError, validate_zip_bytes
from app.services.auction_item_service import AuctionItemService, calculate_bid_increment


@dataclass
class ParsedRow:
    row_number: int
    data: dict[str, Any]


class AuctionItemImportService:
    """Bulk import service for auction items."""

    def __init__(self, db: AsyncSession, settings: Settings | None = None) -> None:
        self.db = db
        self.settings = settings or get_settings()

    async def preflight(self, event_id: UUID, zip_bytes: bytes) -> ImportReport:
        contents = validate_zip_bytes(zip_bytes)
        parsed_rows = self._parse_workbook(contents.workbook_bytes, contents.workbook_filename)

        external_ids = [
            str(value).strip()
            for row in parsed_rows
            if (value := row.data.get("external_id")) not in (None, "")
        ]
        existing_ids = await self._fetch_existing_external_ids(event_id, external_ids)

        results = self._validate_rows(parsed_rows, contents.image_files, existing_ids)
        return self._build_report(results)

    async def commit(self, event_id: UUID, zip_bytes: bytes, user_id: UUID) -> ImportReport:
        contents = validate_zip_bytes(zip_bytes)
        parsed_rows = self._parse_workbook(contents.workbook_bytes, contents.workbook_filename)
        row_lookup = {row.row_number: row for row in parsed_rows}

        external_ids = [
            str(value).strip()
            for row in parsed_rows
            if (value := row.data.get("external_id")) not in (None, "")
        ]
        existing_ids = await self._fetch_existing_external_ids(event_id, external_ids)

        results = []
        for result in self._validate_rows(parsed_rows, contents.image_files, existing_ids):
            if result.status == ImportRowStatus.ERROR:
                results.append(result)
                continue

            try:
                await self._upsert_auction_item(
                    event_id=event_id,
                    row_number=result.row_number,
                    external_id=result.external_id or "",
                    row=row_lookup[result.row_number].data,
                    image_files=contents.image_files,
                    user_id=user_id,
                )
                results.append(result)
            except Exception as exc:  # pragma: no cover - defensive
                await self.db.rollback()
                results.append(
                    ImportRowResult(
                        row_number=result.row_number,
                        external_id=result.external_id,
                        title=result.title,
                        status=ImportRowStatus.ERROR,
                        message=str(exc),
                        image_status=result.image_status,
                        image_count=result.image_count,
                    )
                )

        return self._build_report(results)

    def _parse_workbook(self, workbook_bytes: bytes, workbook_filename: str) -> list[ParsedRow]:
        if workbook_filename.lower().endswith(".csv"):
            return self._parse_csv(workbook_bytes)

        workbook = load_workbook(io.BytesIO(workbook_bytes), read_only=True, data_only=True)
        sheet = workbook.active
        row_iter = sheet.iter_rows(values_only=True)
        header_row = next(row_iter, None)
        if not header_row:
            raise ImportZipValidationError("Workbook is empty")

        headers = [self._normalize_header(value) for value in header_row]
        self._validate_required_headers(headers)

        parsed_rows: list[ParsedRow] = []
        for index, values in enumerate(row_iter, start=2):
            if self._is_empty_row(values):
                continue
            if len(parsed_rows) >= MAX_IMPORT_ROWS:
                raise ImportZipValidationError("Workbook exceeds maximum row limit")

            row_data = {
                headers[i]: values[i] if i < len(values) else None for i in range(len(headers))
            }
            parsed_rows.append(ParsedRow(row_number=index, data=row_data))

        if not parsed_rows:
            raise ImportZipValidationError("Workbook contains no data rows")

        return parsed_rows

    def _parse_csv(self, workbook_bytes: bytes) -> list[ParsedRow]:
        try:
            content = workbook_bytes.decode("utf-8-sig")
        except UnicodeDecodeError as exc:
            raise ImportZipValidationError("CSV must be UTF-8 encoded") from exc

        reader = csv.reader(io.StringIO(content))
        header_row = next(reader, None)
        if not header_row:
            raise ImportZipValidationError("CSV is empty")

        headers = [self._normalize_header(value) for value in header_row]
        self._validate_required_headers(headers)

        parsed_rows: list[ParsedRow] = []
        for index, values in enumerate(reader, start=2):
            if self._is_empty_row(values):
                continue
            if len(parsed_rows) >= MAX_IMPORT_ROWS:
                raise ImportZipValidationError("Workbook exceeds maximum row limit")

            row_data = {
                headers[i]: values[i] if i < len(values) else None for i in range(len(headers))
            }
            parsed_rows.append(ParsedRow(row_number=index, data=row_data))

        if not parsed_rows:
            raise ImportZipValidationError("Workbook contains no data rows")

        return parsed_rows

    def _validate_required_headers(self, headers: list[str]) -> None:
        required_columns = {
            "external_id",
            "title",
            "description",
            "auction_type",
            "fair_market_value",
            "starting_bid",
            "category",
        }
        missing = sorted(required_columns - set(headers))
        if missing:
            raise ImportZipValidationError(f"Missing required columns: {', '.join(missing)}")

        if "image_filename" not in headers and "image_filenames" not in headers:
            raise ImportZipValidationError(
                "Missing required image column: image_filename or image_filenames"
            )

    def _validate_rows(
        self,
        parsed_rows: list[ParsedRow],
        image_files: dict[str, bytes],
        existing_ids: set[str],
    ) -> list[ImportRowResult]:
        allowed_categories = {category.lower(): category for category in ALLOWED_CATEGORIES}
        results: list[ImportRowResult] = []

        external_id_counts = Counter(
            row.data.get("external_id") for row in parsed_rows if row.data.get("external_id")
        )

        for parsed in parsed_rows:
            row_number = parsed.row_number
            data = self._coerce_row(parsed.data)
            external_id = data.get("external_id")
            errors: list[str] = []

            try:
                row = AuctionItemImportRow(**data)
            except Exception as exc:
                results.append(
                    ImportRowResult(
                        row_number=row_number,
                        external_id=external_id,
                        title=data.get("title"),
                        status=ImportRowStatus.ERROR,
                        message=str(exc),
                        image_count=0,
                    )
                )
                continue

            if external_id_counts.get(row.external_id, 0) > 1:
                errors.append("Duplicate external_id in workbook")

            category_key = row.category.strip().lower()
            if category_key not in allowed_categories:
                errors.append("Category is not in the allowed list")

            if row.starting_bid > row.fair_market_value:
                errors.append("Starting bid exceeds fair market value")

            image_filenames = self._get_image_filenames(row)
            image_count = len([name for name in image_filenames if name in image_files])
            image_status = (
                ImportImageStatus.OK
                if image_count == len(image_filenames)
                else ImportImageStatus.MISSING
            )
            if not image_filenames:
                image_status = ImportImageStatus.MISSING
                errors.append("No image filenames provided")
            elif image_count < len(image_filenames):
                errors.append("One or more image files not found in ZIP")

            if errors:
                results.append(
                    ImportRowResult(
                        row_number=row_number,
                        external_id=row.external_id,
                        title=row.title,
                        status=ImportRowStatus.ERROR,
                        message="; ".join(errors),
                        image_status=image_status,
                        image_count=image_count,
                    )
                )
                continue

            status = (
                ImportRowStatus.UPDATED
                if row.external_id in existing_ids
                else ImportRowStatus.CREATED
            )
            results.append(
                ImportRowResult(
                    row_number=row_number,
                    external_id=row.external_id,
                    title=row.title,
                    status=status,
                    message="Ready",
                    image_status=image_status,
                    image_count=image_count,
                )
            )

        return results

    async def _fetch_existing_external_ids(
        self, event_id: UUID, external_ids: list[str]
    ) -> set[str]:
        if not external_ids:
            return set()

        query = select(AuctionItem.external_id).where(
            AuctionItem.event_id == event_id,
            AuctionItem.external_id.in_(external_ids),
        )
        result = await self.db.execute(query)
        return {row[0] for row in result.fetchall()}

    async def _upsert_auction_item(
        self,
        event_id: UUID,
        row_number: int,
        external_id: str,
        row: dict[str, Any],
        image_files: dict[str, bytes],
        user_id: UUID,
    ) -> None:
        row_data = AuctionItemImportRow(**self._coerce_row(row))
        image_filenames = self._get_image_filenames(row_data)

        existing_query = select(AuctionItem).where(
            AuctionItem.event_id == event_id,
            AuctionItem.external_id == external_id,
        )
        result = await self.db.execute(existing_query)
        existing_item = result.scalar_one_or_none()

        if existing_item:
            self._apply_row_to_item(existing_item, row_data)
            await self.db.commit()
            await self.db.refresh(existing_item)
            await self._attach_images(existing_item, image_filenames, image_files)
            return

        service = AuctionItemService(self.db)
        bid_number = await service._get_next_bid_number(event_id)
        bid_increment = calculate_bid_increment(row_data.starting_bid)

        new_item = AuctionItem(
            event_id=event_id,
            external_id=row_data.external_id,
            bid_number=bid_number,
            title=row_data.title,
            description=row_data.description,
            auction_type=AuctionType(row_data.auction_type),
            category=row_data.category,
            starting_bid=row_data.starting_bid,
            bid_increment=bid_increment,
            donor_value=row_data.fair_market_value,
            buy_now_price=row_data.buy_it_now,
            buy_now_enabled=row_data.buy_it_now is not None,
            quantity_available=row_data.quantity or 1,
            donated_by=row_data.donor_name,
            display_priority=row_data.sort_order,
            status=ItemStatus.DRAFT,
            created_by=user_id,
        )

        self.db.add(new_item)
        await self.db.commit()
        await self.db.refresh(new_item)
        await self._attach_images(new_item, image_filenames, image_files)

    def _apply_row_to_item(self, item: AuctionItem, row: AuctionItemImportRow) -> None:
        item.title = row.title
        item.description = row.description
        item.auction_type = AuctionType(row.auction_type)
        item.category = row.category
        item.starting_bid = row.starting_bid
        item.donor_value = row.fair_market_value
        item.buy_now_price = row.buy_it_now
        item.buy_now_enabled = row.buy_it_now is not None
        item.quantity_available = row.quantity or 1
        item.donated_by = row.donor_name
        item.display_priority = row.sort_order

    async def _attach_images(
        self,
        item: AuctionItem,
        image_filenames: list[str],
        image_files: dict[str, bytes],
    ) -> None:
        display_order = 0
        for image_filename in image_filenames:
            if image_filename not in image_files:
                continue

            image_bytes = image_files[image_filename]
            mime_type = magic.from_buffer(image_bytes, mime=True)
            file_path = self._store_image(item.id, image_filename, image_bytes, mime_type)

            media = AuctionItemMedia(
                auction_item_id=item.id,
                media_type=MediaType.IMAGE.value,
                file_path=file_path,
                file_name=image_filename,
                file_size=len(image_bytes),
                mime_type=mime_type,
                display_order=display_order,
            )
            self.db.add(media)
            display_order += 1
        await self.db.commit()

    def _store_image(self, item_id: UUID, filename: str, content: bytes, mime_type: str) -> str:
        if (
            self.settings.azure_storage_connection_string
            and self.settings.azure_storage_account_name
        ):
            from azure.storage.blob import BlobServiceClient, ContentSettings

            blob_service = BlobServiceClient.from_connection_string(
                self.settings.azure_storage_connection_string
            )
            blob_name = f"auction-items/{item_id}/{filename}"
            blob_client = blob_service.get_blob_client(
                container=self.settings.azure_storage_container_name,
                blob=blob_name,
            )
            blob_client.upload_blob(
                content,
                overwrite=True,
                content_settings=ContentSettings(content_type=mime_type),
            )
            return str(blob_client.url)

        from pathlib import Path

        base_dir = Path("static/uploads/auction-items") / str(item_id)
        base_dir.mkdir(parents=True, exist_ok=True)
        file_path = base_dir / filename
        file_path.write_bytes(content)
        return str(file_path)

    def _coerce_row(self, data: dict[str, Any]) -> dict[str, Any]:
        def _to_decimal(value: Any) -> Decimal | None:
            if value is None or value == "":
                return None
            return Decimal(str(value))

        def _to_int(value: Any) -> int | None:
            if value is None or value == "":
                return None
            return int(value)

        def _to_bool(value: Any) -> bool | None:
            if value is None or value == "":
                return None
            if isinstance(value, bool):
                return value
            if isinstance(value, (int, float)):
                return bool(value)
            return str(value).strip().lower() in {"true", "yes", "1"}

        def _to_str(value: Any) -> str | None:
            if value is None:
                return None
            return str(value).strip()

        return {
            "external_id": _to_str(data.get("external_id")),
            "title": _to_str(data.get("title")),
            "description": _to_str(data.get("description")),
            "auction_type": _to_str(data.get("auction_type")),
            "category": _to_str(data.get("category")),
            "starting_bid": _to_decimal(data.get("starting_bid")),
            "fair_market_value": _to_decimal(data.get("fair_market_value")),
            "buy_it_now": _to_decimal(data.get("buy_it_now")),
            "quantity": _to_int(data.get("quantity")),
            "donor_name": _to_str(data.get("donor_name")),
            "tags": _to_str(data.get("tags")),
            "restrictions": _to_str(data.get("restrictions")),
            "fulfillment_notes": _to_str(data.get("fulfillment_notes")),
            "is_featured": _to_bool(data.get("is_featured")),
            "sort_order": _to_int(data.get("sort_order")),
            "image_filename": _to_str(data.get("image_filename")),
            "image_filenames": _to_str(data.get("image_filenames")),
        }

    @staticmethod
    def _get_image_filenames(row: AuctionItemImportRow) -> list[str]:
        if row.image_filenames:
            raw = row.image_filenames
        else:
            raw = row.image_filename or ""

        if not raw:
            return []

        parts = [part.strip() for part in raw.replace(";", ",").replace("\n", ",").split(",")]
        return [part for part in parts if part]

    def _build_report(self, rows: list[ImportRowResult]) -> ImportReport:
        created_count = sum(1 for row in rows if row.status == ImportRowStatus.CREATED)
        updated_count = sum(1 for row in rows if row.status == ImportRowStatus.UPDATED)
        skipped_count = sum(1 for row in rows if row.status == ImportRowStatus.SKIPPED)
        error_count = sum(1 for row in rows if row.status == ImportRowStatus.ERROR)

        report = ImportReport(
            total_rows=len(rows),
            created_count=created_count,
            updated_count=updated_count,
            skipped_count=skipped_count,
            error_count=error_count,
            warnings_count=0,
            rows=rows,
            error_report_url=self._build_error_report_url(rows),
        )
        return report

    def _build_error_report_url(self, rows: list[ImportRowResult]) -> str | None:
        error_rows = [row for row in rows if row.status == ImportRowStatus.ERROR]
        if not error_rows:
            return None

        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(
            [
                "row_number",
                "external_id",
                "title",
                "status",
                "message",
                "image_status",
                "image_count",
            ]
        )
        for row in error_rows:
            writer.writerow(
                [
                    row.row_number,
                    row.external_id or "",
                    row.title or "",
                    row.status.value,
                    row.message,
                    row.image_status.value if row.image_status else "",
                    row.image_count,
                ]
            )

        csv_data = output.getvalue().encode("utf-8")
        encoded = base64.b64encode(csv_data).decode("utf-8")
        return f"data:text/csv;base64,{encoded}"

    @staticmethod
    def _normalize_header(value: Any) -> str:
        if value is None:
            return ""
        return str(value).strip().lower().replace(" ", "_")

    @staticmethod
    def _is_empty_row(values: list[Any] | tuple[Any, ...]) -> bool:
        return all(value is None or str(value).strip() == "" for value in values)
